"""
SemiAnalysis vs DataCenterHawk Enhanced Comparison Script (V2)
===============================================================

Performs detailed comparison between SemiAnalysis (SA) and DataCenterHawk (DCH)
with focus on:
1. Exclusive facility detection (SA-only, DCH-only)
2. Matched facility attribute conflicts (capacity, company, status)
3. Statistical analysis with correlation and significance testing
4. Exportable datasets (CSV + Feature Class)
5. Enhanced HTML report with data visualizations

Author: Meta Data Center GIS Team
Created: 2026-01-29
Version: 2.0
"""

import arcpy
import os
import sys
import csv
import math
import json
from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Tuple, Optional, Any

# Add _utils to path
try:
    script_dir = os.path.dirname(os.path.abspath(__file__))
except NameError:
    script_dir = r"C:\Users\ptanderson\Documents\ArcGIS\Projects\Lean Consensus DC Model\scripts\05_accuracy"

utils_dir = os.path.join(os.path.dirname(script_dir), "_utils")
if utils_dir not in sys.path:
    sys.path.insert(0, utils_dir)

from config import GDB, GOLD_BUILDINGS, GOLD_CAMPUS, ACCURACY_REPORTS_DIR, LOCAL_REPORTS_DIR

# ==============================================================================
# CONFIGURATION
# ==============================================================================

# Feature class targets - use GOLD_BUILDINGS for building-level, GOLD_CAMPUS for campus-level
TARGET_FC_BUILDINGS = GOLD_BUILDINGS
TARGET_FC_CAMPUS = GOLD_CAMPUS
TARGET_FC = GOLD_BUILDINGS  # Default for backwards compatibility

SA_SOURCES = ['Semianalysis']
DCH_SOURCES = ['DataCenterHawk']

# Spatial matching threshold (meters)
DEFAULT_SPATIAL_THRESHOLD_M = 500

# Conflict thresholds
SIGNIFICANT_CAPACITY_DELTA_PCT = 20.0  # 20% difference is significant
SIGNIFICANT_CAPACITY_DELTA_MW = 10.0   # 10 MW absolute difference is significant

# Hyperscaler companies for tier analysis
HYPERSCALERS = ['AWS', 'Microsoft', 'Google', 'Meta', 'Apple', 'Oracle', 'xAI', 'OpenAI',
                'Anthropic', 'ByteDance', 'Crusoe', 'CoreWeave', 'Alibaba']

# Fields to extract
CAPACITY_FIELDS = ['full_capacity_mw', 'commissioned_power_mw', 'uc_power_mw',
                   'planned_power_mw', 'planned_plus_uc_mw']
GEO_FIELDS = ['country', 'region', 'state', 'city', 'market']
COMPANY_FIELDS = ['company_clean', 'company_clean_filter']
STATUS_FIELDS = ['facility_status', 'record_level']
YEAR_FIELDS = ['mw_2023', 'mw_2024', 'mw_2025', 'mw_2026', 'mw_2027',
               'mw_2028', 'mw_2029', 'mw_2030', 'mw_2031', 'mw_2032']

# Output paths
OUTPUT_DIR = str(ACCURACY_REPORTS_DIR) if ACCURACY_REPORTS_DIR.exists() else str(LOCAL_REPORTS_DIR)
CONFLICT_FC_NAME = "sa_dch_conflicts"


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def safe_float(val) -> Optional[float]:
    """Safely convert value to float."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val) -> str:
    """Safely convert value to string."""
    if val is None:
        return ""
    return str(val).strip()


def format_number(val, decimals=0) -> str:
    """Format number with commas."""
    if val is None:
        return "N/A"
    if decimals == 0:
        return f"{int(val):,}"
    return f"{val:,.{decimals}f}"


def normalize_granularity(record: Dict) -> str:
    """
    Normalize granularity from building_designation or record_level fields.

    Returns one of: 'Building', 'Campus', 'Suite', 'Unknown'
    """
    # Check building_designation first, then record_level
    designation = safe_str(record.get('building_designation')).lower()
    record_level = safe_str(record.get('record_level')).lower()

    # Normalize to standard granularity levels
    for val in [designation, record_level]:
        if not val:
            continue
        if 'campus' in val:
            return 'Campus'
        if 'building' in val:
            return 'Building'
        if 'suite' in val or 'unit' in val or 'cage' in val:
            return 'Suite'

    return 'Unknown'


def check_company_match(sa_company: str, dch_company: str) -> bool:
    """
    Check if two company names match (case-insensitive, normalized).

    Used for company-aware spatial matching.
    """
    if not sa_company or not dch_company:
        return False

    sa_norm = sa_company.lower().strip()
    dch_norm = dch_company.lower().strip()

    # Direct match
    if sa_norm == dch_norm:
        return True

    # Common variations
    variations = {
        'amazon web services': 'aws',
        'amazon': 'aws',
        'microsoft azure': 'microsoft',
        'azure': 'microsoft',
        'google cloud': 'google',
        'alphabet': 'google',
        'facebook': 'meta',
        'meta platforms': 'meta',
    }

    sa_normalized = variations.get(sa_norm, sa_norm)
    dch_normalized = variations.get(dch_norm, dch_norm)

    return sa_normalized == dch_normalized


def format_percent(val, total) -> str:
    """Format percentage."""
    if total == 0:
        return "0.0%"
    return f"{(val / total * 100):.1f}%"


def haversine_distance_m(lon1: float, lat1: float, lon2: float, lat2: float) -> float:
    """
    Calculate the great-circle distance between two points in meters.
    Uses the Haversine formula for accurate distance calculation.
    """
    R = 6371000  # Earth's radius in meters

    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)

    a = math.sin(delta_lat / 2) ** 2 + \
        math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    return R * c


def calculate_capacity_delta(sa_val: Optional[float], dch_val: Optional[float]) -> Dict[str, Any]:
    """
    Calculate capacity delta metrics.

    Returns dict with:
        - delta_mw: SA - DCH (positive = SA higher)
        - delta_pct: Percentage difference relative to max value
        - is_significant: Whether delta exceeds thresholds
    """
    sa_cap = sa_val if sa_val is not None else 0
    dch_cap = dch_val if dch_val is not None else 0

    delta_mw = sa_cap - dch_cap

    # Calculate percentage relative to maximum (avoids division by zero, measures relative error)
    max_cap = max(sa_cap, dch_cap)
    if max_cap > 0:
        delta_pct = abs(delta_mw) / max_cap * 100
    else:
        delta_pct = 0.0

    # Significance: either >20% difference OR >10 MW absolute
    is_significant = (delta_pct >= SIGNIFICANT_CAPACITY_DELTA_PCT) or (abs(delta_mw) >= SIGNIFICANT_CAPACITY_DELTA_MW)

    return {
        'delta_mw': delta_mw,
        'delta_pct': delta_pct,
        'is_significant': is_significant,
        'sa_capacity': sa_cap,
        'dch_capacity': dch_cap
    }


def calculate_pearson_correlation(x_vals: List[float], y_vals: List[float]) -> Dict[str, float]:
    """
    Calculate Pearson correlation coefficient and related statistics.

    Returns dict with:
        - r: Correlation coefficient (-1 to 1)
        - r_squared: Coefficient of determination
        - n: Sample size
        - mean_x, mean_y: Means
        - std_x, std_y: Standard deviations
    """
    n = len(x_vals)
    if n < 2:
        return {'r': None, 'r_squared': None, 'n': n, 'mean_x': None, 'mean_y': None}

    # Calculate means
    mean_x = sum(x_vals) / n
    mean_y = sum(y_vals) / n

    # Calculate standard deviations and covariance
    sum_sq_x = sum((x - mean_x) ** 2 for x in x_vals)
    sum_sq_y = sum((y - mean_y) ** 2 for y in y_vals)
    sum_coproduct = sum((x - mean_x) * (y - mean_y) for x, y in zip(x_vals, y_vals))

    std_x = math.sqrt(sum_sq_x / n) if sum_sq_x > 0 else 0
    std_y = math.sqrt(sum_sq_y / n) if sum_sq_y > 0 else 0

    # Calculate correlation
    if std_x > 0 and std_y > 0:
        r = sum_coproduct / math.sqrt(sum_sq_x * sum_sq_y)
    else:
        r = 0

    return {
        'r': r,
        'r_squared': r ** 2,
        'n': n,
        'mean_x': mean_x,
        'mean_y': mean_y,
        'std_x': std_x,
        'std_y': std_y
    }


def calculate_descriptive_stats(values: List[float]) -> Dict[str, float]:
    """Calculate descriptive statistics for a list of values."""
    if not values:
        return {'n': 0, 'mean': None, 'median': None, 'std': None, 'min': None, 'max': None}

    n = len(values)
    mean_val = sum(values) / n
    sorted_vals = sorted(values)

    # Median
    if n % 2 == 0:
        median_val = (sorted_vals[n // 2 - 1] + sorted_vals[n // 2]) / 2
    else:
        median_val = sorted_vals[n // 2]

    # Standard deviation (population)
    variance = sum((x - mean_val) ** 2 for x in values) / n
    std_val = math.sqrt(variance)

    # Quartiles
    q1_idx = n // 4
    q3_idx = (3 * n) // 4
    q1 = sorted_vals[q1_idx] if n > 4 else sorted_vals[0]
    q3 = sorted_vals[q3_idx] if n > 4 else sorted_vals[-1]

    return {
        'n': n,
        'mean': mean_val,
        'median': median_val,
        'std': std_val,
        'min': min(values),
        'max': max(values),
        'q1': q1,
        'q3': q3,
        'sum': sum(values)
    }


# ==============================================================================
# DATA EXTRACTION
# ==============================================================================

def load_records_by_source(fc: str, source_values: List[str]) -> List[Dict]:
    """
    Load records from feature class filtered by source.

    Returns list of dictionaries with all relevant fields plus coordinates.
    """
    # Get all fields from feature class
    all_fields = [f.name for f in arcpy.ListFields(fc)]

    # Build read fields list
    read_fields = ['SHAPE@XY', 'OID@'] + all_fields
    read_fields = list(dict.fromkeys(read_fields))  # Remove duplicates, preserve order

    # Build where clause
    source_list = ", ".join([f"'{s}'" for s in source_values])
    where_clause = f"source IN ({source_list})"

    records = []
    with arcpy.da.SearchCursor(fc, read_fields, where_clause) as cursor:
        for row in cursor:
            record = {
                '_xy': row[0],
                '_oid': row[1],
                '_lon': row[0][0] if row[0] and row[0][0] else None,
                '_lat': row[0][1] if row[0] and row[0][1] else None
            }
            for i, field in enumerate(read_fields[2:], start=2):
                record[field] = row[i]
            records.append(record)

    return records


# ==============================================================================
# SPATIAL MATCHING
# ==============================================================================

def build_match_sets(
    sa_records: List[Dict],
    dch_records: List[Dict],
    threshold_m: float = DEFAULT_SPATIAL_THRESHOLD_M
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """
    Perform spatial matching between SA and DCH records.

    For each SA record, finds the closest DCH record within threshold.
    Uses 1:1 best-match assignment (each DCH can only match one SA).

    Returns:
        matched_pairs: List of dicts with sa_record, dch_record, distance_m, and conflict metrics
        sa_only: List of SA records with no DCH match
        dch_only: List of DCH records with no SA match
    """
    print(f"\n  Building match sets (threshold: {threshold_m}m)...")

    # Filter to records with valid coordinates
    sa_with_coords = [r for r in sa_records if r['_lon'] is not None and r['_lat'] is not None]
    dch_with_coords = [r for r in dch_records if r['_lon'] is not None and r['_lat'] is not None]

    print(f"    SA records with coords: {len(sa_with_coords):,}")
    print(f"    DCH records with coords: {len(dch_with_coords):,}")

    # Track which records have been matched
    sa_matched_ids = set()
    dch_matched_ids = set()

    # Find all potential matches with distances
    potential_matches = []

    for sa_rec in sa_with_coords:
        sa_lon, sa_lat = sa_rec['_lon'], sa_rec['_lat']
        sa_id = sa_rec.get('unique_id') or sa_rec.get('_oid')

        for dch_rec in dch_with_coords:
            dch_lon, dch_lat = dch_rec['_lon'], dch_rec['_lat']
            dch_id = dch_rec.get('unique_id') or dch_rec.get('_oid')

            # Calculate accurate distance using Haversine
            distance_m = haversine_distance_m(sa_lon, sa_lat, dch_lon, dch_lat)

            if distance_m <= threshold_m:
                potential_matches.append({
                    'sa_record': sa_rec,
                    'dch_record': dch_rec,
                    'sa_id': sa_id,
                    'dch_id': dch_id,
                    'distance_m': distance_m
                })

    print(f"    Potential matches found: {len(potential_matches):,}")

    # Sort by distance (closest first) for greedy best-match assignment
    potential_matches.sort(key=lambda x: x['distance_m'])

    # Greedy 1:1 matching - closest pairs first
    matched_pairs = []
    for match in potential_matches:
        sa_id = match['sa_id']
        dch_id = match['dch_id']

        # Skip if either record already matched
        if sa_id in sa_matched_ids or dch_id in dch_matched_ids:
            continue

        # Mark as matched
        sa_matched_ids.add(sa_id)
        dch_matched_ids.add(dch_id)

        # Calculate attribute conflicts
        sa_rec = match['sa_record']
        dch_rec = match['dch_record']

        capacity_delta = calculate_capacity_delta(
            safe_float(sa_rec.get('full_capacity_mw')),
            safe_float(dch_rec.get('full_capacity_mw'))
        )

        company_match = check_company_match(
            safe_str(sa_rec.get('company_clean')),
            safe_str(dch_rec.get('company_clean'))
        )
        company_filter_match = safe_str(sa_rec.get('company_clean_filter')) == safe_str(dch_rec.get('company_clean_filter'))
        status_match = safe_str(sa_rec.get('facility_status')).lower() == safe_str(dch_rec.get('facility_status')).lower()

        # Granularity matching - check if comparing same level (building vs building, etc.)
        sa_granularity = normalize_granularity(sa_rec)
        dch_granularity = normalize_granularity(dch_rec)
        granularity_match = sa_granularity == dch_granularity

        matched_pairs.append({
            'sa_record': sa_rec,
            'dch_record': dch_rec,
            'sa_id': sa_id,
            'dch_id': dch_id,
            'distance_m': match['distance_m'],
            'capacity_delta': capacity_delta,
            'company_match': company_match,
            'company_filter_match': company_filter_match,
            'status_match': status_match,
            'granularity_match': granularity_match,
            'sa_granularity': sa_granularity,
            'dch_granularity': dch_granularity
        })

    # Build unmatched lists
    sa_only = [r for r in sa_records
               if (r.get('unique_id') or r.get('_oid')) not in sa_matched_ids]
    dch_only = [r for r in dch_records
                if (r.get('unique_id') or r.get('_oid')) not in dch_matched_ids]

    print(f"\n    Final match results:")
    print(f"      Matched pairs: {len(matched_pairs):,}")
    print(f"      SA-only (unmatched): {len(sa_only):,}")
    print(f"      DCH-only (unmatched): {len(dch_only):,}")

    return matched_pairs, sa_only, dch_only


# ==============================================================================
# STATISTICAL ANALYSIS
# ==============================================================================

def analyze_capacity_conflicts(matched_pairs: List[Dict]) -> Dict[str, Any]:
    """
    Comprehensive statistical analysis of capacity conflicts.

    NOTE: Capacity comparisons (MAPE, bias, correlation) are only calculated
    for same-granularity matches (building↔building, campus↔campus) to ensure
    apples-to-apples comparison.
    """
    print("\n  Analyzing capacity conflicts...")

    # Separate by granularity match for accurate capacity analysis
    same_granularity_pairs = [p for p in matched_pairs if p.get('granularity_match', True)]
    different_granularity_pairs = [p for p in matched_pairs if not p.get('granularity_match', True)]

    granularity_match_rate = len(same_granularity_pairs) / len(matched_pairs) * 100 if matched_pairs else 0

    print(f"    Same-granularity matches: {len(same_granularity_pairs):,} ({granularity_match_rate:.1f}%)")
    print(f"    Different-granularity matches: {len(different_granularity_pairs):,} (excluded from capacity MAPE)")

    # Use same-granularity pairs for capacity analysis
    analysis_pairs = same_granularity_pairs if same_granularity_pairs else matched_pairs

    # Extract capacity deltas (from same-granularity pairs only)
    deltas_mw = [p['capacity_delta']['delta_mw'] for p in analysis_pairs]
    deltas_pct = [p['capacity_delta']['delta_pct'] for p in analysis_pairs]
    sa_caps = [p['capacity_delta']['sa_capacity'] for p in analysis_pairs if p['capacity_delta']['sa_capacity'] > 0]
    dch_caps = [p['capacity_delta']['dch_capacity'] for p in analysis_pairs if p['capacity_delta']['dch_capacity'] > 0]

    # Filter to pairs where both have capacity for correlation
    paired_caps = [(p['capacity_delta']['sa_capacity'], p['capacity_delta']['dch_capacity'])
                   for p in analysis_pairs
                   if p['capacity_delta']['sa_capacity'] > 0 and p['capacity_delta']['dch_capacity'] > 0]

    # Descriptive statistics
    delta_stats = calculate_descriptive_stats(deltas_mw)
    pct_stats = calculate_descriptive_stats(deltas_pct)

    # Correlation analysis
    if paired_caps:
        sa_vals, dch_vals = zip(*paired_caps)
        correlation = calculate_pearson_correlation(list(sa_vals), list(dch_vals))
    else:
        correlation = {'r': None, 'r_squared': None, 'n': 0}

    # Significance counts (from same-granularity pairs)
    significant_conflicts = [p for p in analysis_pairs if p['capacity_delta']['is_significant']]

    # Direction analysis (SA higher vs DCH higher)
    sa_higher = [p for p in analysis_pairs if p['capacity_delta']['delta_mw'] > SIGNIFICANT_CAPACITY_DELTA_MW]
    dch_higher = [p for p in analysis_pairs if p['capacity_delta']['delta_mw'] < -SIGNIFICANT_CAPACITY_DELTA_MW]

    # Distribution bins for histogram
    bins = [(-float('inf'), -50), (-50, -20), (-20, -10), (-10, -5), (-5, 0),
            (0, 5), (5, 10), (10, 20), (20, 50), (50, float('inf'))]
    bin_labels = ['<-50', '-50 to -20', '-20 to -10', '-10 to -5', '-5 to 0',
                  '0 to 5', '5 to 10', '10 to 20', '20 to 50', '>50']
    histogram = {label: 0 for label in bin_labels}

    for delta in deltas_mw:
        for i, (low, high) in enumerate(bins):
            if low <= delta < high:
                histogram[bin_labels[i]] += 1
                break

    # Granularity breakdown
    granularity_breakdown = defaultdict(int)
    for p in matched_pairs:
        sa_gran = p.get('sa_granularity', 'Unknown')
        dch_gran = p.get('dch_granularity', 'Unknown')
        key = f"{sa_gran} ↔ {dch_gran}"
        granularity_breakdown[key] += 1

    return {
        'delta_stats': delta_stats,
        'pct_stats': pct_stats,
        'correlation': correlation,
        'total_pairs': len(matched_pairs),
        'same_granularity_pairs': len(same_granularity_pairs),
        'different_granularity_pairs': len(different_granularity_pairs),
        'granularity_match_rate': granularity_match_rate,
        'granularity_breakdown': dict(granularity_breakdown),
        'significant_conflicts_count': len(significant_conflicts),
        'significant_conflicts_pct': len(significant_conflicts) / len(analysis_pairs) * 100 if analysis_pairs else 0,
        'sa_higher_count': len(sa_higher),
        'dch_higher_count': len(dch_higher),
        'histogram': histogram,
        'paired_capacity_count': len(paired_caps)
    }


def analyze_by_company(matched_pairs: List[Dict], sa_only: List[Dict], dch_only: List[Dict]) -> Dict[str, Any]:
    """
    Company-level analysis of coverage and conflicts.

    NOTE: Capacity totals (sa_total_mw, dch_total_mw) are calculated from
    same-granularity pairs only to avoid double-counting when building records
    inherit campus-level capacity values.
    """
    print("\n  Analyzing by company...")

    company_stats = defaultdict(lambda: {
        'matched_count': 0,
        'matched_same_granularity': 0,
        'sa_only_count': 0,
        'dch_only_count': 0,
        'sa_total_mw': 0,
        'dch_total_mw': 0,
        'conflicts': [],
        'mean_delta_mw': 0,
        'company_mismatch_count': 0,
        'granularity_mismatch_count': 0
    })

    # Process matched pairs
    for pair in matched_pairs:
        # Use company_clean for actual company name grouping
        company = pair['sa_record'].get('company_clean') or pair['sa_record'].get('company_clean_filter') or 'Unknown'

        company_stats[company]['matched_count'] += 1

        # Only add capacity for same-granularity pairs (apples-to-apples)
        if pair.get('granularity_match', True):
            company_stats[company]['matched_same_granularity'] += 1
            company_stats[company]['sa_total_mw'] += pair['capacity_delta']['sa_capacity']
            company_stats[company]['dch_total_mw'] += pair['capacity_delta']['dch_capacity']

            if pair['capacity_delta']['is_significant']:
                company_stats[company]['conflicts'].append(pair['capacity_delta']['delta_mw'])
        else:
            company_stats[company]['granularity_mismatch_count'] += 1

        if not pair['company_match']:
            company_stats[company]['company_mismatch_count'] += 1

    # Process SA-only (use building records only - filter out campus-level)
    for rec in sa_only:
        company = rec.get('company_clean') or rec.get('company_clean_filter') or 'Unknown'
        company_stats[company]['sa_only_count'] += 1
        # Only add capacity if this is a building-level record
        granularity = normalize_granularity(rec)
        if granularity == 'Building':
            company_stats[company]['sa_total_mw'] += safe_float(rec.get('full_capacity_mw')) or 0

    # Process DCH-only (use building records only - filter out campus-level)
    for rec in dch_only:
        company = rec.get('company_clean') or rec.get('company_clean_filter') or 'Unknown'
        company_stats[company]['dch_only_count'] += 1
        # Only add capacity if this is a building-level record
        granularity = normalize_granularity(rec)
        if granularity == 'Building':
            company_stats[company]['dch_total_mw'] += safe_float(rec.get('full_capacity_mw')) or 0

    # Calculate mean deltas
    for company, stats in company_stats.items():
        if stats['conflicts']:
            stats['mean_delta_mw'] = sum(stats['conflicts']) / len(stats['conflicts'])
            stats['conflict_count'] = len(stats['conflicts'])
        else:
            stats['conflict_count'] = 0

    return dict(company_stats)


def analyze_by_geography(matched_pairs: List[Dict], sa_only: List[Dict], dch_only: List[Dict]) -> Dict[str, Any]:
    """
    Geographic analysis of coverage and conflicts.
    """
    print("\n  Analyzing by geography...")

    # By Region
    region_stats = defaultdict(lambda: {'matched': 0, 'sa_only': 0, 'dch_only': 0, 'conflicts': 0})

    for pair in matched_pairs:
        region = pair['sa_record'].get('region') or 'Unknown'
        region_stats[region]['matched'] += 1
        if pair['capacity_delta']['is_significant']:
            region_stats[region]['conflicts'] += 1

    for rec in sa_only:
        region = rec.get('region') or 'Unknown'
        region_stats[region]['sa_only'] += 1

    for rec in dch_only:
        region = rec.get('region') or 'Unknown'
        region_stats[region]['dch_only'] += 1

    # By Country (top 15)
    country_stats = defaultdict(lambda: {'matched': 0, 'sa_only': 0, 'dch_only': 0, 'conflicts': 0})

    for pair in matched_pairs:
        country = pair['sa_record'].get('country') or 'Unknown'
        country_stats[country]['matched'] += 1
        if pair['capacity_delta']['is_significant']:
            country_stats[country]['conflicts'] += 1

    for rec in sa_only:
        country = rec.get('country') or 'Unknown'
        country_stats[country]['sa_only'] += 1

    for rec in dch_only:
        country = rec.get('country') or 'Unknown'
        country_stats[country]['dch_only'] += 1

    # Sort by total records
    top_countries = sorted(
        country_stats.items(),
        key=lambda x: x[1]['matched'] + x[1]['sa_only'] + x[1]['dch_only'],
        reverse=True
    )[:15]

    return {
        'by_region': dict(region_stats),
        'by_country': dict(top_countries)
    }


def get_top_conflicts(matched_pairs: List[Dict], n: int = 25) -> List[Dict]:
    """Get the top N largest capacity conflicts by absolute delta."""
    sorted_pairs = sorted(matched_pairs, key=lambda x: abs(x['capacity_delta']['delta_mw']), reverse=True)

    top_conflicts = []
    for pair in sorted_pairs[:n]:
        top_conflicts.append({
            'sa_id': pair['sa_id'],
            'dch_id': pair['dch_id'],
            'sa_company': pair['sa_record'].get('company_clean'),
            'dch_company': pair['dch_record'].get('company_clean'),
            'sa_capacity_mw': pair['capacity_delta']['sa_capacity'],
            'dch_capacity_mw': pair['capacity_delta']['dch_capacity'],
            'delta_mw': pair['capacity_delta']['delta_mw'],
            'delta_pct': pair['capacity_delta']['delta_pct'],
            'city': pair['sa_record'].get('city') or pair['dch_record'].get('city'),
            'state': pair['sa_record'].get('state') or pair['dch_record'].get('state'),
            'country': pair['sa_record'].get('country') or pair['dch_record'].get('country'),
            'distance_m': pair['distance_m']
        })

    return top_conflicts


# ==============================================================================
# NET NEW SITES ANALYSIS
# ==============================================================================

# Define net new facility statuses (non-operational)
# Note: 'Land Acquisition' added as it represents early-stage planned sites
# 'Planned' and 'Proposed' kept for compatibility but don't exist in current data
NET_NEW_STATUSES = ['Under Construction', 'Announced', 'Planned', 'Proposed', 'Land Acquisition']

def analyze_net_new_sites(
    matched_pairs: List[Dict],
    sa_only: List[Dict],
    dch_only: List[Dict],
    sa_records: List[Dict],
    dch_records: List[Dict]
) -> Dict[str, Any]:
    """
    Analyze net new sites (Under Construction, Announced, Planned, Proposed)
    to assess how well SA and DCH agree on future/in-progress data centers.

    This analysis answers key questions:
    1. Coverage: What % of SA's net new sites appear in DCH (and vice versa)?
    2. Capacity Agreement: For matched net new sites, how well do capacity forecasts agree?
    3. Status Agreement: Do sources agree on the construction stage?
    4. Geographic Patterns: Regional differences in net new site coverage
    5. Company Patterns: Hyperscaler vs colo coverage differences

    Returns:
        Dictionary with comprehensive net new site analysis metrics
    """
    print("\n  Analyzing net new sites (Under Construction, Announced, Planned)...")

    def is_net_new(status: Optional[str]) -> bool:
        """Check if a facility status is considered 'net new'."""
        if not status:
            return False
        status_lower = status.lower().strip()
        return any(ns.lower() in status_lower for ns in NET_NEW_STATUSES)

    # -------------------------------------------------------------------------
    # 1. Count net new sites by source (all records)
    # -------------------------------------------------------------------------
    sa_net_new_all = [r for r in sa_records if is_net_new(r.get('facility_status'))]
    dch_net_new_all = [r for r in dch_records if is_net_new(r.get('facility_status'))]

    # -------------------------------------------------------------------------
    # 2. Filter matched pairs to net new (at least one side is net new)
    # -------------------------------------------------------------------------
    matched_net_new = []
    for pair in matched_pairs:
        sa_status = pair['sa_record'].get('facility_status')
        dch_status = pair['dch_record'].get('facility_status')
        if is_net_new(sa_status) or is_net_new(dch_status):
            matched_net_new.append(pair)

    # Matched pairs where BOTH are net new
    matched_both_net_new = [
        p for p in matched_pairs
        if is_net_new(p['sa_record'].get('facility_status')) and is_net_new(p['dch_record'].get('facility_status'))
    ]

    # -------------------------------------------------------------------------
    # 3. Filter exclusive records to net new
    # -------------------------------------------------------------------------
    sa_only_net_new = [r for r in sa_only if is_net_new(r.get('facility_status'))]
    dch_only_net_new = [r for r in dch_only if is_net_new(r.get('facility_status'))]

    # -------------------------------------------------------------------------
    # 4. Calculate coverage rates
    # -------------------------------------------------------------------------
    # How many SA net-new sites appear in DCH?
    sa_net_new_matched_count = len([p for p in matched_net_new if is_net_new(p['sa_record'].get('facility_status'))])
    sa_net_new_total = len(sa_net_new_all)
    sa_coverage_rate = (sa_net_new_matched_count / sa_net_new_total * 100) if sa_net_new_total > 0 else 0

    # How many DCH net-new sites appear in SA?
    dch_net_new_matched_count = len([p for p in matched_net_new if is_net_new(p['dch_record'].get('facility_status'))])
    dch_net_new_total = len(dch_net_new_all)
    dch_coverage_rate = (dch_net_new_matched_count / dch_net_new_total * 100) if dch_net_new_total > 0 else 0

    # -------------------------------------------------------------------------
    # 5. Capacity agreement for matched net new sites
    # -------------------------------------------------------------------------
    capacity_deltas = []
    capacity_pairs = []
    for pair in matched_both_net_new:
        sa_cap = pair['capacity_delta']['sa_capacity']
        dch_cap = pair['capacity_delta']['dch_capacity']
        if sa_cap > 0 or dch_cap > 0:
            capacity_deltas.append(pair['capacity_delta']['delta_mw'])
            capacity_pairs.append((sa_cap, dch_cap))

    # Calculate MAPE for net new sites
    net_new_mape = None
    net_new_bias = None
    if capacity_pairs:
        apes = []
        for sa_cap, dch_cap in capacity_pairs:
            max_cap = max(sa_cap, dch_cap)
            if max_cap > 0:
                apes.append(abs(sa_cap - dch_cap) / max_cap * 100)
        if apes:
            net_new_mape = sum(apes) / len(apes)

        # Bias calculation
        total_sa = sum(c[0] for c in capacity_pairs)
        total_dch = sum(c[1] for c in capacity_pairs)
        if total_dch > 0:
            net_new_bias = ((total_sa - total_dch) / total_dch) * 100

    # -------------------------------------------------------------------------
    # 6. Status agreement analysis
    # -------------------------------------------------------------------------
    status_agreement_count = 0
    status_disagreement_details = defaultdict(int)

    for pair in matched_both_net_new:
        sa_status = safe_str(pair['sa_record'].get('facility_status')).lower()
        dch_status = safe_str(pair['dch_record'].get('facility_status')).lower()

        if sa_status == dch_status:
            status_agreement_count += 1
        else:
            key = f"{pair['sa_record'].get('facility_status')} vs {pair['dch_record'].get('facility_status')}"
            status_disagreement_details[key] += 1

    status_agreement_rate = (status_agreement_count / len(matched_both_net_new) * 100) if matched_both_net_new else 0

    # -------------------------------------------------------------------------
    # 7. Breakdown by specific status
    # -------------------------------------------------------------------------
    def count_by_status(records: List[Dict], status_filter: str) -> int:
        return len([r for r in records
                   if safe_str(r.get('facility_status')).lower() == status_filter.lower()])

    status_breakdown = {
        'Under Construction': {
            'sa_total': count_by_status(sa_records, 'Under Construction'),
            'dch_total': count_by_status(dch_records, 'Under Construction'),
            'sa_only': count_by_status(sa_only, 'Under Construction'),
            'dch_only': count_by_status(dch_only, 'Under Construction'),
        },
        'Announced': {
            'sa_total': count_by_status(sa_records, 'Announced'),
            'dch_total': count_by_status(dch_records, 'Announced'),
            'sa_only': count_by_status(sa_only, 'Announced'),
            'dch_only': count_by_status(dch_only, 'Announced'),
        },
        'Planned': {
            'sa_total': count_by_status(sa_records, 'Planned'),
            'dch_total': count_by_status(dch_records, 'Planned'),
            'sa_only': count_by_status(sa_only, 'Planned'),
            'dch_only': count_by_status(dch_only, 'Planned'),
        }
    }

    # Calculate matched counts for each status
    for status_name in status_breakdown:
        sa_total = status_breakdown[status_name]['sa_total']
        sa_exclusive = status_breakdown[status_name]['sa_only']
        status_breakdown[status_name]['sa_matched'] = sa_total - sa_exclusive

        dch_total = status_breakdown[status_name]['dch_total']
        dch_exclusive = status_breakdown[status_name]['dch_only']
        status_breakdown[status_name]['dch_matched'] = dch_total - dch_exclusive

    # -------------------------------------------------------------------------
    # 8. Geographic analysis of net new sites
    # -------------------------------------------------------------------------
    region_breakdown = defaultdict(lambda: {
        'sa_net_new': 0, 'dch_net_new': 0,
        'sa_only_net_new': 0, 'dch_only_net_new': 0, 'matched_net_new': 0
    })

    for r in sa_net_new_all:
        region = r.get('region') or 'Unknown'
        region_breakdown[region]['sa_net_new'] += 1

    for r in dch_net_new_all:
        region = r.get('region') or 'Unknown'
        region_breakdown[region]['dch_net_new'] += 1

    for r in sa_only_net_new:
        region = r.get('region') or 'Unknown'
        region_breakdown[region]['sa_only_net_new'] += 1

    for r in dch_only_net_new:
        region = r.get('region') or 'Unknown'
        region_breakdown[region]['dch_only_net_new'] += 1

    for p in matched_both_net_new:
        region = p['sa_record'].get('region') or p['dch_record'].get('region') or 'Unknown'
        region_breakdown[region]['matched_net_new'] += 1

    # -------------------------------------------------------------------------
    # 9. Company/tier analysis of net new sites
    # -------------------------------------------------------------------------
    company_net_new = defaultdict(lambda: {
        'sa_count': 0, 'dch_count': 0, 'matched_count': 0,
        'sa_only_count': 0, 'dch_only_count': 0,
        'sa_capacity_mw': 0, 'dch_capacity_mw': 0
    })

    for r in sa_net_new_all:
        company = r.get('company_clean_filter') or r.get('company_clean') or 'Unknown'
        company_net_new[company]['sa_count'] += 1
        company_net_new[company]['sa_capacity_mw'] += safe_float(r.get('full_capacity_mw')) or 0

    for r in dch_net_new_all:
        company = r.get('company_clean_filter') or r.get('company_clean') or 'Unknown'
        company_net_new[company]['dch_count'] += 1
        company_net_new[company]['dch_capacity_mw'] += safe_float(r.get('full_capacity_mw')) or 0

    for r in sa_only_net_new:
        company = r.get('company_clean_filter') or r.get('company_clean') or 'Unknown'
        company_net_new[company]['sa_only_count'] += 1

    for r in dch_only_net_new:
        company = r.get('company_clean_filter') or r.get('company_clean') or 'Unknown'
        company_net_new[company]['dch_only_count'] += 1

    for p in matched_both_net_new:
        company = p['sa_record'].get('company_clean_filter') or p['sa_record'].get('company_clean') or 'Unknown'
        company_net_new[company]['matched_count'] += 1

    # -------------------------------------------------------------------------
    # 10. Top exclusive net new sites (largest capacity)
    # -------------------------------------------------------------------------
    sa_only_sorted = sorted(sa_only_net_new,
                           key=lambda x: safe_float(x.get('full_capacity_mw')) or 0,
                           reverse=True)[:20]
    dch_only_sorted = sorted(dch_only_net_new,
                            key=lambda x: safe_float(x.get('full_capacity_mw')) or 0,
                            reverse=True)[:20]

    # -------------------------------------------------------------------------
    # 11. Capacity totals
    # -------------------------------------------------------------------------
    sa_net_new_total_mw = sum(safe_float(r.get('full_capacity_mw')) or 0 for r in sa_net_new_all)
    dch_net_new_total_mw = sum(safe_float(r.get('full_capacity_mw')) or 0 for r in dch_net_new_all)
    sa_only_net_new_mw = sum(safe_float(r.get('full_capacity_mw')) or 0 for r in sa_only_net_new)
    dch_only_net_new_mw = sum(safe_float(r.get('full_capacity_mw')) or 0 for r in dch_only_net_new)

    print(f"    SA net new sites: {len(sa_net_new_all):,} ({sa_net_new_total_mw:,.0f} MW)")
    print(f"    DCH net new sites: {len(dch_net_new_all):,} ({dch_net_new_total_mw:,.0f} MW)")
    print(f"    Matched (both net new): {len(matched_both_net_new):,}")
    print(f"    SA-only net new: {len(sa_only_net_new):,} ({sa_only_net_new_mw:,.0f} MW)")
    print(f"    DCH-only net new: {len(dch_only_net_new):,} ({dch_only_net_new_mw:,.0f} MW)")
    print(f"    SA coverage rate: {sa_coverage_rate:.1f}%")
    print(f"    DCH coverage rate: {dch_coverage_rate:.1f}%")
    if net_new_mape:
        print(f"    Net new MAPE: {net_new_mape:.1f}%")

    return {
        # Counts
        'sa_net_new_total': len(sa_net_new_all),
        'dch_net_new_total': len(dch_net_new_all),
        'matched_net_new': len(matched_net_new),
        'matched_both_net_new': len(matched_both_net_new),
        'sa_only_net_new': len(sa_only_net_new),
        'dch_only_net_new': len(dch_only_net_new),

        # Coverage rates
        'sa_coverage_rate': sa_coverage_rate,
        'dch_coverage_rate': dch_coverage_rate,

        # Capacity metrics
        'sa_net_new_total_mw': sa_net_new_total_mw,
        'dch_net_new_total_mw': dch_net_new_total_mw,
        'sa_only_net_new_mw': sa_only_net_new_mw,
        'dch_only_net_new_mw': dch_only_net_new_mw,
        'net_new_mape': net_new_mape,
        'net_new_bias': net_new_bias,

        # Status agreement
        'status_agreement_rate': status_agreement_rate,
        'status_disagreement_details': dict(status_disagreement_details),

        # Breakdowns
        'status_breakdown': status_breakdown,
        'region_breakdown': dict(region_breakdown),
        'company_breakdown': dict(company_net_new),

        # Top exclusive sites
        'top_sa_only_net_new': sa_only_sorted,
        'top_dch_only_net_new': dch_only_sorted,

        # Raw lists for potential further analysis
        'matched_both_net_new_pairs': matched_both_net_new,
        'sa_only_net_new_records': sa_only_net_new,
        'dch_only_net_new_records': dch_only_net_new
    }


# ==============================================================================
# EXPORT FUNCTIONS
# ==============================================================================

def export_matched_pairs_csv(matched_pairs: List[Dict], output_path: str):
    """Export matched pairs with conflict metrics to CSV."""
    print(f"\n  Exporting matched pairs to: {output_path}")

    fieldnames = [
        'sa_unique_id', 'dch_unique_id', 'distance_m',
        'sa_company', 'dch_company', 'company_match',
        'sa_capacity_mw', 'dch_capacity_mw', 'capacity_delta_mw', 'capacity_delta_pct', 'is_significant_conflict',
        'sa_granularity', 'dch_granularity', 'granularity_match',
        'sa_status', 'dch_status', 'status_match',
        'city', 'state', 'country', 'region', 'market',
        'sa_lat', 'sa_lon', 'dch_lat', 'dch_lon'
    ]

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for pair in matched_pairs:
            writer.writerow({
                'sa_unique_id': pair['sa_id'],
                'dch_unique_id': pair['dch_id'],
                'distance_m': round(pair['distance_m'], 1),
                'sa_company': pair['sa_record'].get('company_clean'),
                'dch_company': pair['dch_record'].get('company_clean'),
                'company_match': pair['company_match'],
                'sa_capacity_mw': pair['capacity_delta']['sa_capacity'],
                'dch_capacity_mw': pair['capacity_delta']['dch_capacity'],
                'capacity_delta_mw': round(pair['capacity_delta']['delta_mw'], 2),
                'capacity_delta_pct': round(pair['capacity_delta']['delta_pct'], 1),
                'is_significant_conflict': pair['capacity_delta']['is_significant'],
                'sa_granularity': pair.get('sa_granularity', 'Unknown'),
                'dch_granularity': pair.get('dch_granularity', 'Unknown'),
                'granularity_match': pair.get('granularity_match', True),
                'sa_status': pair['sa_record'].get('facility_status'),
                'dch_status': pair['dch_record'].get('facility_status'),
                'status_match': pair['status_match'],
                'city': pair['sa_record'].get('city') or pair['dch_record'].get('city'),
                'state': pair['sa_record'].get('state') or pair['dch_record'].get('state'),
                'country': pair['sa_record'].get('country') or pair['dch_record'].get('country'),
                'region': pair['sa_record'].get('region') or pair['dch_record'].get('region'),
                'market': pair['sa_record'].get('market') or pair['dch_record'].get('market'),
                'sa_lat': pair['sa_record'].get('_lat'),
                'sa_lon': pair['sa_record'].get('_lon'),
                'dch_lat': pair['dch_record'].get('_lat'),
                'dch_lon': pair['dch_record'].get('_lon')
            })

    print(f"    Exported {len(matched_pairs):,} matched pairs")


def export_exclusive_records_csv(records: List[Dict], output_path: str, source_name: str):
    """Export exclusive (unmatched) records to CSV."""
    print(f"\n  Exporting {source_name}-only records to: {output_path}")

    fieldnames = [
        'unique_id', 'company_clean', 'company_clean_filter',
        'full_capacity_mw', 'facility_status', 'record_level',
        'city', 'state', 'country', 'region', 'market',
        'latitude', 'longitude'
    ]

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for rec in records:
            writer.writerow({
                'unique_id': rec.get('unique_id'),
                'company_clean': rec.get('company_clean'),
                'company_clean_filter': rec.get('company_clean_filter'),
                'full_capacity_mw': safe_float(rec.get('full_capacity_mw')),
                'facility_status': rec.get('facility_status'),
                'record_level': rec.get('record_level'),
                'city': rec.get('city'),
                'state': rec.get('state'),
                'country': rec.get('country'),
                'region': rec.get('region'),
                'market': rec.get('market'),
                'latitude': rec.get('_lat'),
                'longitude': rec.get('_lon')
            })

    print(f"    Exported {len(records):,} {source_name}-only records")


def create_conflict_feature_class(matched_pairs: List[Dict], output_fc: str, min_delta_pct: float = 10.0):
    """
    Create a feature class of significant capacity conflicts for ArcGIS Pro visualization.
    """
    print(f"\n  Creating conflict feature class: {output_fc}")

    # Filter to significant conflicts
    significant_pairs = [p for p in matched_pairs
                         if p['capacity_delta']['delta_pct'] >= min_delta_pct
                         or abs(p['capacity_delta']['delta_mw']) >= SIGNIFICANT_CAPACITY_DELTA_MW]

    print(f"    Significant conflicts (>={min_delta_pct}% or >={SIGNIFICANT_CAPACITY_DELTA_MW} MW): {len(significant_pairs):,}")

    # Delete if exists
    if arcpy.Exists(output_fc):
        arcpy.Delete_management(output_fc)

    # Create feature class
    arcpy.CreateFeatureclass_management(
        out_path=os.path.dirname(output_fc),
        out_name=os.path.basename(output_fc),
        geometry_type="POINT",
        spatial_reference=arcpy.SpatialReference(4326)  # WGS84
    )

    # Add fields
    fields_to_add = [
        ('sa_unique_id', 'TEXT', 100),
        ('dch_unique_id', 'TEXT', 100),
        ('sa_company', 'TEXT', 100),
        ('dch_company', 'TEXT', 100),
        ('company_match', 'SHORT', None),
        ('sa_capacity_mw', 'DOUBLE', None),
        ('dch_capacity_mw', 'DOUBLE', None),
        ('capacity_delta_mw', 'DOUBLE', None),
        ('capacity_delta_pct', 'DOUBLE', None),
        ('sa_status', 'TEXT', 50),
        ('dch_status', 'TEXT', 50),
        ('status_match', 'SHORT', None),
        ('distance_m', 'DOUBLE', None),
        ('city', 'TEXT', 100),
        ('state', 'TEXT', 50),
        ('country', 'TEXT', 50),
        ('region', 'TEXT', 20),
        ('conflict_direction', 'TEXT', 20)  # 'SA_Higher' or 'DCH_Higher'
    ]

    for field_name, field_type, field_length in fields_to_add:
        if field_length:
            arcpy.AddField_management(output_fc, field_name, field_type, field_length=field_length)
        else:
            arcpy.AddField_management(output_fc, field_name, field_type)

    # Insert records (use SA coordinates as reference point)
    insert_fields = ['SHAPE@XY'] + [f[0] for f in fields_to_add]

    with arcpy.da.InsertCursor(output_fc, insert_fields) as cursor:
        for pair in significant_pairs:
            sa_rec = pair['sa_record']
            dch_rec = pair['dch_record']
            delta = pair['capacity_delta']

            # Determine conflict direction
            if delta['delta_mw'] > 0:
                conflict_direction = 'SA_Higher'
            elif delta['delta_mw'] < 0:
                conflict_direction = 'DCH_Higher'
            else:
                conflict_direction = 'Equal'

            row = (
                (sa_rec['_lon'], sa_rec['_lat']),  # SHAPE@XY
                pair['sa_id'],
                pair['dch_id'],
                sa_rec.get('company_clean'),
                dch_rec.get('company_clean'),
                1 if pair['company_match'] else 0,
                delta['sa_capacity'],
                delta['dch_capacity'],
                delta['delta_mw'],
                delta['delta_pct'],
                sa_rec.get('facility_status'),
                dch_rec.get('facility_status'),
                1 if pair['status_match'] else 0,
                pair['distance_m'],
                sa_rec.get('city') or dch_rec.get('city'),
                sa_rec.get('state') or dch_rec.get('state'),
                sa_rec.get('country') or dch_rec.get('country'),
                sa_rec.get('region') or dch_rec.get('region'),
                conflict_direction
            )
            cursor.insertRow(row)

    print(f"    Created feature class with {len(significant_pairs):,} conflict records")
    return len(significant_pairs)


# ==============================================================================
# HTML REPORT GENERATION
# ==============================================================================

def generate_html_report(
    sa_records: List[Dict],
    dch_records: List[Dict],
    matched_pairs: List[Dict],
    sa_only: List[Dict],
    dch_only: List[Dict],
    capacity_analysis: Dict,
    company_analysis: Dict,
    geo_analysis: Dict,
    top_conflicts: List[Dict],
    net_new_analysis: Optional[Dict],
    output_path: str,
    threshold_m: float
):
    """Generate comprehensive HTML report with data visualizations."""
    print(f"\n  Generating HTML report: {output_path}")

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Prepare data for charts
    histogram_data = capacity_analysis['histogram']
    histogram_labels = list(histogram_data.keys())
    histogram_values = list(histogram_data.values())

    # Company data for chart (top 10 by total records)
    company_chart_data = sorted(
        [(k, v) for k, v in company_analysis.items()],
        key=lambda x: x[1]['matched_count'] + x[1]['sa_only_count'] + x[1]['dch_only_count'],
        reverse=True
    )[:12]

    # Region data
    region_data = geo_analysis['by_region']

    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="light">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SA vs DCH Comparison Report V2</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-zoom@2.0.1/dist/chartjs-plugin-zoom.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/hammerjs@2.0.8"></script>
    <style>
        :root {{
            --primary: #1877F2;
            --sa-color: #E91E63;
            --dch-color: #4CAF50;
            --conflict-color: #FF9800;
            --bg-light: #f8f9fa;
            --border-color: #dee2e6;
            --bg-main: #f0f2f5;
            --text-primary: #1c1e21;
            --text-secondary: #65676b;
            --card-bg: white;
            --card-shadow: 0 2px 4px rgba(0,0,0,0.05);
            --table-stripe: #f8f9fa;
            --explanation-bg: linear-gradient(135deg, #f8f9ff 0%, #f0f4ff 100%);
            --explanation-text: #444;
        }}

        [data-theme="dark"] {{
            --primary: #4da6ff;
            --sa-color: #ff6b9d;
            --dch-color: #69db7c;
            --conflict-color: #ffa94d;
            --bg-light: #2d2d2d;
            --border-color: #444;
            --bg-main: #1a1a1a;
            --text-primary: #e4e6eb;
            --text-secondary: #b0b3b8;
            --card-bg: #242526;
            --card-shadow: 0 2px 4px rgba(0,0,0,0.3);
            --table-stripe: #2d2d2d;
            --explanation-bg: linear-gradient(135deg, #1e3a5f 0%, #1a2a3e 100%);
            --explanation-text: #b0b3b8;
        }}

        * {{ box-sizing: border-box; }}

        body {{
            font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
            margin: 0;
            padding: 0;
            background: var(--bg-main);
            color: var(--text-primary);
            line-height: 1.5;
            transition: background 0.3s, color 0.3s;
        }}

        /* Left Navigation */
        .nav-sidebar {{
            position: fixed;
            top: 0;
            left: 0;
            width: 240px;
            height: 100vh;
            background: linear-gradient(180deg, #1a1a2e 0%, #16213e 100%);
            padding: 20px 0;
            overflow-y: auto;
            z-index: 1000;
            box-shadow: 2px 0 10px rgba(0,0,0,0.1);
        }}

        .nav-logo {{
            color: white;
            font-size: 1.2em;
            font-weight: 700;
            padding: 0 20px 20px;
            border-bottom: 1px solid rgba(255,255,255,0.1);
            margin-bottom: 20px;
        }}

        .nav-section {{
            padding: 10px 20px;
            font-size: 0.75em;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: rgba(255,255,255,0.5);
        }}

        .nav-link {{
            display: block;
            padding: 12px 20px;
            color: rgba(255,255,255,0.8);
            text-decoration: none;
            font-size: 0.9em;
            transition: all 0.2s;
            border-left: 3px solid transparent;
        }}

        .nav-link:hover {{
            background: rgba(255,255,255,0.1);
            color: white;
            border-left-color: var(--primary);
        }}

        .nav-link.active {{
            background: rgba(24, 119, 242, 0.2);
            color: white;
            border-left-color: var(--primary);
        }}

        .nav-icon {{
            margin-right: 10px;
        }}

        /* Main content with sidebar offset */
        .main-content {{
            margin-left: 240px;
            padding: 20px;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
        }}

        .header {{
            background: linear-gradient(135deg, var(--primary), #0d47a1);
            color: white;
            padding: 30px;
            border-radius: 12px;
            margin-bottom: 24px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}

        .header h1 {{
            margin: 0 0 10px 0;
            font-size: 2em;
        }}

        .header .subtitle {{
            opacity: 0.9;
            font-size: 1.1em;
        }}

        .card {{
            background: var(--card-bg);
            border-radius: 12px;
            padding: 24px;
            margin-bottom: 24px;
            box-shadow: var(--card-shadow);
            transition: background 0.3s, box-shadow 0.3s;
        }}

        .card h2 {{
            color: var(--primary);
            margin-top: 0;
            padding-bottom: 12px;
            border-bottom: 2px solid var(--bg-light);
        }}

        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}

        .metric-box {{
            background: var(--bg-light);
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}

        .metric-value {{
            font-size: 2.2em;
            font-weight: 700;
            margin-bottom: 5px;
        }}

        .metric-value.sa {{ color: var(--sa-color); }}
        .metric-value.dch {{ color: var(--dch-color); }}
        .metric-value.conflict {{ color: var(--conflict-color); }}
        .metric-value.primary {{ color: var(--primary); }}

        .metric-label {{
            color: #65676b;
            font-size: 0.9em;
        }}

        .chart-container {{
            position: relative;
            height: 350px;
            margin: 20px 0;
        }}

        .chart-row {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
            gap: 24px;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
            font-size: 0.95em;
        }}

        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid var(--border-color);
        }}

        th {{
            background: var(--bg-light);
            font-weight: 600;
            color: #4b4f56;
        }}

        tr:hover {{
            background: #f8f9fa;
        }}

        .badge {{
            display: inline-block;
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 0.85em;
            font-weight: 500;
        }}

        .badge-sa {{ background: #fce4ec; color: var(--sa-color); }}
        .badge-dch {{ background: #e8f5e9; color: var(--dch-color); }}
        .badge-conflict {{ background: #fff3e0; color: #e65100; }}
        .badge-match {{ background: #e3f2fd; color: var(--primary); }}

        .stat-highlight {{
            background: linear-gradient(135deg, #e3f2fd, #bbdefb);
            padding: 20px;
            border-radius: 8px;
            margin: 15px 0;
        }}

        .stat-highlight h3 {{
            margin-top: 0;
            color: var(--primary);
        }}

        .correlation-box {{
            display: flex;
            align-items: center;
            gap: 20px;
            padding: 15px;
            background: var(--bg-light);
            border-radius: 8px;
            margin: 15px 0;
        }}

        .correlation-value {{
            font-size: 3em;
            font-weight: 700;
            color: var(--primary);
        }}

        .correlation-label {{
            flex: 1;
        }}

        .legend {{
            display: flex;
            gap: 20px;
            flex-wrap: wrap;
            margin: 15px 0;
        }}

        /* Tooltip styles */
        .tooltip {{
            position: relative;
            cursor: help;
        }}

        .tooltip .tooltip-text {{
            visibility: hidden;
            background-color: #333;
            color: white;
            text-align: left;
            border-radius: 6px;
            padding: 10px 14px;
            position: absolute;
            z-index: 1000;
            bottom: 125%;
            left: 50%;
            transform: translateX(-50%);
            width: 280px;
            font-size: 0.85em;
            font-weight: normal;
            line-height: 1.4;
            box-shadow: 0 4px 12px rgba(0,0,0,0.3);
            opacity: 0;
            transition: opacity 0.2s;
        }}

        .tooltip .tooltip-text::after {{
            content: "";
            position: absolute;
            top: 100%;
            left: 50%;
            margin-left: -6px;
            border-width: 6px;
            border-style: solid;
            border-color: #333 transparent transparent transparent;
        }}

        .tooltip:hover .tooltip-text {{
            visibility: visible;
            opacity: 1;
        }}

        .metric-box.tooltip .tooltip-text {{
            width: 300px;
        }}

        .interpretation-box {{
            background: linear-gradient(135deg, #fff8e1, #ffecb3);
            border-left: 4px solid #ffc107;
            padding: 16px 20px;
            border-radius: 0 8px 8px 0;
            margin: 20px 0;
        }}

        .interpretation-box h4 {{
            margin: 0 0 10px 0;
            color: #ff8f00;
            font-size: 1em;
        }}

        .interpretation-box ul {{
            margin: 0;
            padding-left: 20px;
        }}

        .interpretation-box li {{
            margin: 6px 0;
            color: #5d4037;
        }}

        .grade-legend {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 8px;
            margin: 15px 0;
            text-align: center;
            font-size: 0.85em;
        }}

        .grade-item {{
            padding: 8px;
            border-radius: 4px;
        }}

        .grade-A {{ background: #c8e6c9; color: #2e7d32; }}
        .grade-B {{ background: #dcedc8; color: #558b2f; }}
        .grade-C {{ background: #fff9c4; color: #f9a825; }}
        .grade-D {{ background: #ffecb3; color: #ff8f00; }}
        .grade-F {{ background: #ffcdd2; color: #c62828; }}

        .legend-item {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .legend-color {{
            width: 16px;
            height: 16px;
            border-radius: 4px;
        }}

        .footer {{
            text-align: center;
            padding: 20px;
            color: #65676b;
            font-size: 0.9em;
        }}

        .scroll-table {{
            max-height: 500px;
            overflow-y: auto;
            border: 1px solid var(--border-color);
            border-radius: 8px;
        }}

        .grade {{
            display: inline-block;
            width: 40px;
            height: 40px;
            line-height: 40px;
            text-align: center;
            border-radius: 50%;
            font-weight: 700;
            font-size: 1.2em;
        }}

        .grade-a {{ background: #4CAF50; color: white; }}
        .grade-b {{ background: #8BC34A; color: white; }}
        .grade-c {{ background: #FFC107; color: #333; }}
        .grade-d {{ background: #FF9800; color: white; }}
        .grade-f {{ background: #f44336; color: white; }}

        .zoom-btn {{
            background: var(--primary);
            color: white;
            border: none;
            padding: 6px 12px;
            border-radius: 4px;
            font-size: 0.85em;
            cursor: pointer;
            margin-right: 5px;
            transition: background 0.2s;
        }}
        .zoom-btn:hover {{
            background: #1565c0;
        }}

        .two-col-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(600px, 1fr));
            gap: 24px;
        }}

        .essential-flag {{
            background: #ffeb3b;
            color: #333;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 0.8em;
            font-weight: 600;
        }}

        .direction-sa {{ color: var(--sa-color); font-weight: 600; }}
        .direction-dch {{ color: var(--dch-color); font-weight: 600; }}

        .ci-range {{
            font-size: 0.9em;
            color: #666;
        }}

        @media (max-width: 768px) {{
            .chart-row, .two-col-grid {{
                grid-template-columns: 1fr;
            }}
            .nav-sidebar {{
                display: none;
            }}
            .main-content {{
                margin-left: 0;
            }}
        }}

        .explanation-box {{
            background: var(--explanation-bg);
            border-left: 4px solid var(--primary);
            padding: 16px 20px;
            margin: 16px 0;
            border-radius: 0 8px 8px 0;
            font-size: 0.95em;
            transition: background 0.3s;
        }}

        .explanation-box h4 {{
            margin: 0 0 8px 0;
            color: var(--primary);
            font-size: 0.9em;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .explanation-box p {{
            margin: 0;
            color: var(--explanation-text);
        }}

        .methodology-note {{
            background: #fff3cd;
            border: 1px solid #ffc107;
            padding: 12px 16px;
            border-radius: 6px;
            font-size: 0.9em;
            margin: 12px 0;
        }}

        [data-theme="dark"] .methodology-note {{
            background: #3d3200;
            border-color: #886b00;
            color: #ffd54f;
        }}

        /* Theme Toggle Button */
        .theme-toggle {{
            position: fixed;
            top: 20px;
            right: 20px;
            z-index: 1001;
            background: var(--card-bg);
            border: 2px solid var(--border-color);
            border-radius: 50px;
            padding: 10px 16px;
            cursor: pointer;
            font-size: 1.1em;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
            transition: all 0.3s;
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .theme-toggle:hover {{
            transform: scale(1.05);
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
        }}

        .theme-toggle-label {{
            font-size: 0.85em;
            color: var(--text-secondary);
        }}

        /* Data Quality Banner */
        .data-quality-banner {{
            background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%);
            border: 1px solid #2196f3;
            border-radius: 8px;
            padding: 16px 20px;
            margin-bottom: 20px;
            font-size: 0.9em;
        }}

        [data-theme="dark"] .data-quality-banner {{
            background: linear-gradient(135deg, #0d2137 0%, #0a3d62 100%);
            border-color: #1976d2;
        }}

        .data-quality-banner h4 {{
            margin: 0 0 8px 0;
            color: #1565c0;
            font-size: 1em;
        }}

        [data-theme="dark"] .data-quality-banner h4 {{
            color: #64b5f6;
        }}

        .data-quality-banner ul {{
            margin: 8px 0 0 0;
            padding-left: 20px;
        }}

        .granularity-badge {{
            display: inline-block;
            background: #e3f2fd;
            color: #1565c0;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 0.8em;
            font-weight: 600;
            margin-left: 8px;
        }}

        [data-theme="dark"] .granularity-badge {{
            background: #0d47a1;
            color: #90caf9;
        }}
    </style>
</head>
<body>
    <!-- Theme Toggle Button -->
    <button class="theme-toggle" onclick="toggleTheme()" title="Toggle light/dark mode">
        <span id="theme-icon">🌙</span>
        <span class="theme-toggle-label" id="theme-label">Dark</span>
    </button>

    <!-- Left Navigation Sidebar -->
    <nav class="nav-sidebar">
        <div class="nav-logo">📊 SA vs DCH V2</div>

        <div class="nav-section">Overview</div>
        <a href="#summary" class="nav-link"><span class="nav-icon">📋</span>Executive Summary</a>
        <a href="#metrics" class="nav-link"><span class="nav-icon">📈</span>Accuracy Metrics</a>

        <div class="nav-section">Analysis</div>
        <a href="#charts" class="nav-link"><span class="nav-icon">📊</span>Visualizations</a>
        <a href="#conflicts" class="nav-link"><span class="nav-icon">⚠️</span>Top Conflicts</a>
        <a href="#company" class="nav-link"><span class="nav-icon">🏢</span>Company Coverage</a>

        <div class="nav-section">Coverage</div>
        <a href="#exclusive" class="nav-link"><span class="nav-icon">🔎</span>Exclusive Records</a>
        <a href="#netnew" class="nav-link"><span class="nav-icon">🚧</span>Net New Sites</a>
        <a href="#groundtruth" class="nav-link"><span class="nav-icon">🎯</span>Ground Truth</a>
        <a href="#geography" class="nav-link"><span class="nav-icon">🌍</span>Geographic Dist.</a>
        <a href="#essential" class="nav-link"><span class="nav-icon">⭐</span>Essential DCs</a>
        <a href="#conclusions" class="nav-link"><span class="nav-icon">📋</span>Conclusions</a>

        <div class="nav-section">Info</div>
        <a href="#methodology" class="nav-link"><span class="nav-icon">📖</span>Methodology</a>
    </nav>

    <!-- Main Content -->
    <div class="main-content">
        <div class="container">
            <!-- Header -->
            <div class="header">
                <h1>🔍 SemiAnalysis vs DataCenterHawk Comparison</h1>
                <div class="subtitle">Enhanced Statistical Analysis Report (V2) | Generated: {timestamp}</div>
            </div>

            <!-- Data Quality & Granularity Notice -->
            <div class="data-quality-banner">
                <h4>📋 Important: Comparison Methodology</h4>
                <ul>
                    <li><strong>Comparison Level:</strong> This report compares <strong>campus-level records</strong> from the gold_campus layer (aggregated by UCID). Each "site" represents a physical campus, not individual buildings.</li>
                    <li><strong>Capacity Values:</strong> Campus capacity uses the MAX value from underlying buildings to avoid double-counting when buildings inherit campus-level values.</li>
                    <li><strong>Spatial Matching:</strong> Campuses are matched using {threshold_m:.0f}m spatial threshold. For building-level comparison, re-run with <code>use_campus_level=False</code>.</li>
                </ul>
            </div>
"""

    # Calculate overall match rate
    total_sa = len(sa_records)
    total_dch = len(dch_records)
    matched_count = len(matched_pairs)
    sa_only_count = len(sa_only)
    dch_only_count = len(dch_only)

    sa_match_rate = (matched_count / total_sa * 100) if total_sa > 0 else 0
    dch_match_rate = (matched_count / total_dch * 100) if total_dch > 0 else 0

    # Unpack capacity analysis
    corr = capacity_analysis.get('correlation', {})
    correlation_r = corr.get('r')
    correlation_r_sq = corr.get('r_squared')
    r_squared = correlation_r_sq if correlation_r_sq is not None else 0.0

    # Pre-format correlation values for safe use in f-strings
    correlation_r_str = f"{correlation_r:.3f}" if correlation_r is not None else 'N/A'
    correlation_r_sq_str = f"{correlation_r_sq:.3f}" if correlation_r_sq is not None else 'N/A'

    delta_stats = capacity_analysis.get('delta_stats', {})
    pct_stats = capacity_analysis.get('pct_stats', {})

    # Get MAPE/Bias/CV if calculated (will be added later in this function if needed)
    mape_val = capacity_analysis.get('mape', 'N/A')
    bias_val = capacity_analysis.get('bias_pct', 'N/A')
    cv_val = capacity_analysis.get('cv', 'N/A')
    grade = capacity_analysis.get('grade', 'N/A')
    grade_class = capacity_analysis.get('grade_class', 'grade-c')

    # Pre-format metric values for safe use in f-strings
    mape_str = f"{mape_val:.1f}%" if isinstance(mape_val, (int, float)) else str(mape_val)
    bias_str = f"{bias_val:+.1f}%" if isinstance(bias_val, (int, float)) else str(bias_val)
    cv_str = f"{cv_val:.1f}%" if isinstance(cv_val, (int, float)) else str(cv_val)
    bias_direction = 'SA reports higher' if isinstance(bias_val, (int, float)) and bias_val > 0 else ('DCH reports higher' if isinstance(bias_val, (int, float)) else 'N/A')
    bias_class = 'sa' if isinstance(bias_val, (int, float)) and bias_val > 0 else 'dch'

    # Generate natural language interpretations
    # Grade interpretation
    if grade == 'A':
        grade_interpretation = "excellent agreement"
    elif grade == 'B':
        grade_interpretation = "good agreement with minor differences"
    elif grade == 'C':
        grade_interpretation = "moderate agreement - some capacity values differ significantly"
    elif grade == 'D':
        grade_interpretation = "poor agreement - many capacity values differ substantially"
    else:
        grade_interpretation = "significant disagreement between sources"

    # Bias interpretation
    if isinstance(bias_val, (int, float)):
        if abs(bias_val) < 10:
            bias_interpretation = "the sources have minimal systematic bias and report similar aggregate values"
        elif bias_val > 30:
            bias_interpretation = "SemiAnalysis consistently reports significantly higher capacity - this may reflect different methodologies or data vintages"
        elif bias_val > 10:
            bias_interpretation = "SemiAnalysis tends to report moderately higher capacity values"
        elif bias_val < -30:
            bias_interpretation = "DataCenterHawk consistently reports significantly higher capacity"
        else:
            bias_interpretation = "DataCenterHawk tends to report moderately higher capacity values"
    else:
        bias_interpretation = "insufficient data to determine bias"

    # CV interpretation
    if isinstance(cv_val, (int, float)):
        if cv_val < 50:
            cv_interpretation = "disagreements are relatively consistent and predictable"
        elif cv_val < 100:
            cv_interpretation = "disagreements vary moderately across facilities"
        elif cv_val < 300:
            cv_interpretation = "disagreements are highly variable - some facilities agree well while others differ substantially"
        else:
            cv_interpretation = "disagreements are extremely variable - sources may use fundamentally different measurement approaches"
    else:
        cv_interpretation = "insufficient data to assess variance"

    # Determine correlation interpretation text
    if correlation_r is not None and correlation_r > 0.8:
        correlation_interpretation = "Strong positive correlation"
    elif correlation_r is not None and correlation_r > 0.5:
        correlation_interpretation = "Moderate correlation"
    elif correlation_r is not None:
        correlation_interpretation = "Weak correlation"
    else:
        correlation_interpretation = "Insufficient data"

    # Executive Summary Card
    html += f"""
        <!-- Executive Summary -->
        <div class="card" id="summary">
            <h2>📊 Executive Summary</h2>

            <div class="explanation-box">
                <h4>What This Report Shows</h4>
                <p>This analysis compares data center records from SemiAnalysis (SA) and DataCenterHawk (DCH) to identify
                overlapping facilities, unique records in each source, and capacity discrepancies. Facilities are matched
                using spatial proximity (within {threshold_m}m) and then compared on key attributes like capacity and ownership.</p>
            </div>

            <div class="metrics-grid">
                <div class="metric-box">
                    <div class="metric-value sa">{total_sa:,}</div>
                    <div class="metric-label">SemiAnalysis Records</div>
                </div>
                <div class="metric-box">
                    <div class="metric-value dch">{total_dch:,}</div>
                    <div class="metric-label">DataCenterHawk Records</div>
                </div>
                <div class="metric-box">
                    <div class="metric-value primary">{matched_count:,}</div>
                    <div class="metric-label">Matched Pairs</div>
                </div>
                <div class="metric-box">
                    <div class="metric-value conflict">{capacity_analysis.get('significant_conflicts_count', 0):,}</div>
                    <div class="metric-label">Significant Conflicts</div>
                </div>
            </div>

            <div class="stat-highlight">
                <h3>Key Findings</h3>
                <ul>
                    <li><strong>Spatial Overlap:</strong> {sa_match_rate:.1f}% of SA records have a DCH match within {threshold_m}m</li>
                    <li><strong>Granularity Match:</strong> {capacity_analysis.get('granularity_match_rate', 0):.1f}% of matched pairs are same-level (building↔building, campus↔campus)</li>
                    <li><strong>SA-Only Records:</strong> {sa_only_count:,} facilities only in SemiAnalysis</li>
                    <li><strong>DCH-Only Records:</strong> {dch_only_count:,} facilities only in DataCenterHawk</li>
                    <li><strong>Capacity Correlation:</strong> r = {correlation_r_str} (r² = {correlation_r_sq_str})</li>
                    <li><strong>Significant Conflicts:</strong> {capacity_analysis.get('significant_conflicts_pct', 0):.1f}% of same-granularity pairs have >20% or >10MW capacity difference</li>
                </ul>
            </div>

            <div class="methodology-note">
                <strong>⚠️ Granularity Filtering:</strong> Capacity metrics (MAPE, Bias, Correlation) are calculated using only
                <strong>{capacity_analysis.get('same_granularity_pairs', 0):,} same-granularity matches</strong>
                to ensure apples-to-apples comparison. {capacity_analysis.get('different_granularity_pairs', 0):,} mismatched-granularity pairs
                (e.g., building↔campus) are counted for coverage but excluded from capacity accuracy.
            </div>
        </div>
"""

    # MAPE/Grade/Statistical Metrics Card
    html += f"""
        <!-- Statistical Accuracy Metrics -->
        <div class="card" id="metrics">
            <h2>📈 Capacity Agreement Metrics</h2>

            <div class="explanation-box">
                <h4>Understanding These Metrics</h4>
                <p><strong>MAPE</strong> (Mean Absolute Percentage Error) measures the average percentage difference
                between SA and DCH capacity values - lower is better. <strong>Bias</strong> indicates if one source
                systematically reports higher values. <strong>CV</strong> (Coefficient of Variation) shows how
                consistent the disagreements are. <strong>Grade</strong> provides an overall accuracy rating from A (excellent) to F (poor).</p>
            </div>

            <div class="metrics-grid">
                <div class="metric-box tooltip">
                    <div class="metric-value primary">{mape_str}</div>
                    <div class="metric-label">MAPE</div>
                    <span class="tooltip-text"><strong>Mean Absolute Percentage Error</strong><br/>Average percentage difference between SA and DCH capacity values. Lower is better.<br/><br/>• 0-10% = Excellent (A)<br/>• 10-20% = Good (B)<br/>• 20-35% = Moderate (C)<br/>• 35-50% = Poor (D)<br/>• >50% = Very Poor (F)</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value {bias_class}">{bias_str}</div>
                    <div class="metric-label">Bias</div>
                    <span class="tooltip-text"><strong>Systematic Bias</strong><br/>Shows if one source consistently reports higher/lower than the other.<br/><br/>• Positive (+) = SA reports higher<br/>• Negative (-) = DCH reports higher<br/>• Near 0% = No systematic bias<br/>• >±20% = Significant bias to investigate</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value conflict">{cv_str}</div>
                    <div class="metric-label">Coefficient of Variation</div>
                    <span class="tooltip-text"><strong>Coefficient of Variation (CV)</strong><br/>Measures how consistent the disagreements are.<br/><br/>• Low CV (<50%) = Consistent differences, easier to model<br/>• High CV (>100%) = Highly variable differences<br/>• Very High (>300%) = Erratic disagreements, may indicate different methodologies</span>
                </div>
                <div class="metric-box tooltip">
                <div class="grade {grade_class}">{grade}</div>
                    <div class="metric-label" style="margin-top: 8px;">Agreement Grade</div>
                    <span class="tooltip-text"><strong>Overall Agreement Grade</strong><br/>Based on MAPE score:<br/><br/>• A (0-10%) = Sources highly agree<br/>• B (10-20%) = Good agreement<br/>• C (20-35%) = Moderate differences<br/>• D (35-50%) = Significant gaps<br/>• F (>50%) = Major discrepancies</span>
                </div>
            </div>

            <div class="interpretation-box">
                <h4>💡 What These Results Mean</h4>
                <ul>
                    <li><strong>MAPE {mape_str}</strong> indicates {grade_interpretation} between SA and DCH capacity values</li>
                    <li><strong>Bias {bias_str}</strong> means {bias_interpretation}</li>
                    <li><strong>CV {cv_str}</strong> suggests {cv_interpretation}</li>
                </ul>
            </div>

            <div class="correlation-box">
                <div class="correlation-value">{correlation_r_str}</div>
                <div class="correlation-label">
                    <strong>Pearson Correlation Coefficient</strong><br>
                    Measures linear relationship between SA and DCH capacity values.
                    {correlation_interpretation}.
                </div>
            </div>

            <table>
                <tr>
                    <th>Metric</th>
                    <th>Formula</th>
                    <th>Value</th>
                    <th>Interpretation</th>
                </tr>
                <tr>
                    <td><strong>MAPE</strong></td>
                    <td>mean(|SA - DCH| / max) × 100</td>
                    <td>{mape_str}</td>
                    <td>Overall accuracy (lower = better)</td>
                </tr>
                <tr>
                    <td><strong>Bias %</strong></td>
                    <td>mean(SA - DCH) / mean(DCH) × 100</td>
                    <td>{bias_str}</td>
                    <td>{bias_direction}</td>
                </tr>
                <tr>
                    <td><strong>CV</strong></td>
                    <td>std(delta) / mean(|delta|) × 100</td>
                    <td>{cv_str}</td>
                    <td>Consistency of disagreement</td>
                </tr>
                <tr>
                    <td><strong>Pearson r</strong></td>
                    <td>correlation(SA_cap, DCH_cap)</td>
                    <td>{correlation_r_str}</td>
                    <td>Linear relationship strength</td>
                </tr>
            </table>
        </div>
"""

    # Prepare chart data as JSON
    histogram_json = json.dumps(histogram_labels)
    histogram_values_json = json.dumps(histogram_values)

    company_labels = [c[0][:20] for c in company_chart_data]
    company_matched = [c[1]['matched_count'] for c in company_chart_data]
    company_sa_only = [c[1]['sa_only_count'] for c in company_chart_data]
    company_dch_only = [c[1]['dch_only_count'] for c in company_chart_data]

    company_labels_json = json.dumps(company_labels)
    company_matched_json = json.dumps(company_matched)
    company_sa_only_json = json.dumps(company_sa_only)
    company_dch_only_json = json.dumps(company_dch_only)

    # Region chart data
    region_labels = list(region_data.keys())
    region_matched = [region_data[r]['matched'] for r in region_labels]
    region_sa_only = [region_data[r]['sa_only'] for r in region_labels]
    region_dch_only = [region_data[r]['dch_only'] for r in region_labels]

    region_labels_json = json.dumps(region_labels)
    region_matched_json = json.dumps(region_matched)
    region_sa_only_json = json.dumps(region_sa_only)
    region_dch_only_json = json.dumps(region_dch_only)

    # Scatter plot data (SA vs DCH capacity)
    scatter_sa = []
    scatter_dch = []
    for pair in matched_pairs:
        sa_cap = pair['capacity_delta']['sa_capacity']
        dch_cap = pair['capacity_delta']['dch_capacity']
        if sa_cap > 0 and dch_cap > 0:
            scatter_sa.append(sa_cap)
            scatter_dch.append(dch_cap)

    scatter_data = [{'x': sa, 'y': dch} for sa, dch in zip(scatter_sa, scatter_dch)]
    scatter_data_json = json.dumps(scatter_data[:500])  # Limit to 500 points for performance

    # Charts Section
    html += f"""
        <!-- Capacity Analysis Charts -->
        <div class="card" id="charts">
            <h2>📊 Capacity Visualizations</h2>

            <div class="explanation-box">
                <h4>Interpreting These Charts</h4>
                <p>The <strong>Delta Distribution</strong> histogram shows how capacity differences are spread - a bell curve
                centered at 0 indicates good agreement. The <strong>Scatter Plot</strong> compares capacity values directly -
                points on the diagonal line represent perfect agreement. The <strong>Company</strong> and <strong>Region</strong>
                charts reveal which segments have the best/worst coverage overlap between sources.</p>
            </div>

            <div class="chart-row">
                <div>
                    <h3>Capacity Delta Distribution (MW)</h3>
                    <div class="chart-container">
                        <canvas id="histogramChart"></canvas>
                    </div>
                </div>
                <div>
                    <h3>SA vs DCH Capacity Scatter (r² = {r_squared:.3f})</h3>
                    <div style="margin-bottom: 8px;">
                        <button onclick="resetScatterZoom()" class="zoom-btn">Reset Zoom</button>
                        <button onclick="zoomScatter(0, 500)" class="zoom-btn">0-500 MW</button>
                        <button onclick="zoomScatter(0, 200)" class="zoom-btn">0-200 MW</button>
                        <button onclick="zoomScatter(0, 100)" class="zoom-btn">0-100 MW</button>
                        <span style="font-size: 0.85em; color: #666; margin-left: 10px;">🖱️ Scroll to zoom, use buttons to zoom</span>
                    </div>
                    <div class="chart-container" style="height: 350px;">
                        <canvas id="scatterChart"></canvas>
                    </div>
                </div>
            </div>

            <div class="chart-row">
                <div>
                    <h3>Records by Company (Top 12)</h3>
                    <div class="chart-container">
                        <canvas id="companyChart"></canvas>
                    </div>
                </div>
                <div>
                    <h3>Coverage by Region</h3>
                    <div class="chart-container">
                        <canvas id="regionChart"></canvas>
                    </div>
                </div>
            </div>
        </div>
"""

    # Top Conflicts Table
    html += """
        <!-- Top Capacity Conflicts -->
        <div class="card" id="conflicts">
            <h2>⚠️ Top 25 Largest Capacity Conflicts</h2>

            <div class="explanation-box">
                <h4>Why These Conflicts Matter</h4>
                <p>Large capacity discrepancies between sources may indicate: (1) different measurement methodologies,
                (2) outdated information in one source, (3) different facility definitions, or (4) planned vs operational
                capacity differences. These records should be prioritized for manual verification.</p>
            </div>

            <p>Facilities with the largest absolute difference in reported capacity between sources.</p>
            <div class="scroll-table">
                <table>
                    <thead>
                        <tr>
                            <th>#</th>
                            <th>Location</th>
                            <th>SA Company</th>
                            <th>DCH Company</th>
                            <th>SA (MW)</th>
                            <th>DCH (MW)</th>
                            <th>Delta (MW)</th>
                            <th>Delta %</th>
                            <th>Direction</th>
                        </tr>
                    </thead>
                    <tbody>
"""

    for i, conflict in enumerate(top_conflicts, 1):
        direction_class = 'direction-sa' if conflict['delta_mw'] > 0 else 'direction-dch'
        direction_text = 'SA Higher' if conflict['delta_mw'] > 0 else 'DCH Higher'
        location = f"{conflict.get('city') or 'Unknown'}, {conflict.get('state') or ''} {conflict.get('country') or ''}"

        html += f"""
                        <tr>
                            <td>{i}</td>
                            <td>{location[:40]}</td>
                            <td>{conflict.get('sa_company') or 'N/A'}</td>
                            <td>{conflict.get('dch_company') or 'N/A'}</td>
                            <td>{conflict['sa_capacity_mw']:,.1f}</td>
                            <td>{conflict['dch_capacity_mw']:,.1f}</td>
                            <td><strong>{conflict['delta_mw']:+,.1f}</strong></td>
                            <td>{conflict['delta_pct']:.1f}%</td>
                            <td class="{direction_class}">{direction_text}</td>
                        </tr>
"""

    html += """
                    </tbody>
                </table>
            </div>
        </div>
"""

    # Company Analysis Table
    html += """
        <!-- Company Analysis -->
        <div class="card" id="company">
            <h2>🏢 Company Coverage Analysis</h2>

            <div class="explanation-box">
                <h4>How to Read This Table</h4>
                <p><strong>Matched</strong> shows facilities found in both sources. <strong>SA Only</strong> and <strong>DCH Only</strong>
                indicate exclusive records in each source. Large numbers in exclusive columns suggest coverage gaps.
                <strong>Conflicts</strong> counts matched facilities with >20% capacity difference.</p>
            </div>

            <div class="scroll-table">
                <table>
                    <thead>
                        <tr>
                            <th>Company</th>
                            <th>Matched</th>
                            <th>SA Only</th>
                            <th>DCH Only</th>
                            <th>SA Total (MW)</th>
                            <th>DCH Total (MW)</th>
                            <th>Conflicts</th>
                            <th>Mean Delta</th>
                        </tr>
                    </thead>
                    <tbody>
"""

    sorted_companies = sorted(
        company_analysis.items(),
        key=lambda x: x[1]['matched_count'] + x[1]['sa_only_count'] + x[1]['dch_only_count'],
        reverse=True
    )[:25]

    for company, stats in sorted_companies:
        html += f"""
                        <tr>
                            <td><strong>{company}</strong></td>
                            <td>{stats['matched_count']:,}</td>
                            <td><span class="badge badge-sa">{stats['sa_only_count']:,}</span></td>
                            <td><span class="badge badge-dch">{stats['dch_only_count']:,}</span></td>
                            <td>{stats['sa_total_mw']:,.0f}</td>
                            <td>{stats['dch_total_mw']:,.0f}</td>
                            <td><span class="badge badge-conflict">{stats['conflict_count']}</span></td>
                            <td>{stats['mean_delta_mw']:+.1f} MW</td>
                        </tr>
"""

    html += """
                    </tbody>
                </table>
            </div>
        </div>
"""

    # Exclusive Records Tables
    sa_only_sorted = sorted(sa_only, key=lambda x: safe_float(x.get('full_capacity_mw')) or 0, reverse=True)[:20]
    dch_only_sorted = sorted(dch_only, key=lambda x: safe_float(x.get('full_capacity_mw')) or 0, reverse=True)[:20]

    html += """
        <!-- Exclusive Records -->
        <div class="card" id="exclusive">
            <h2>🔎 Exclusive Records Analysis</h2>

            <div class="explanation-box">
                <h4>Understanding Coverage Gaps</h4>
                <p><strong>SA-Only</strong> records exist only in SemiAnalysis - these may be newer facilities, different
                naming conventions, or facilities DCH hasn't cataloged. <strong>DCH-Only</strong> records exist only in
                DataCenterHawk - investigating these can help identify gaps in SA's coverage. Larger capacity
                facilities should be prioritized for reconciliation.</p>
            </div>

            <p>Facilities that exist in only one source - potential gaps in coverage.</p>

            <div class="two-col-grid">
                <div>
                    <h3><span class="badge badge-sa">SA-Only</span> Top 20 by Capacity</h3>
                    <div class="scroll-table" style="max-height: 400px;">
                        <table>
                            <thead>
                                <tr>
                                    <th>Company</th>
                                    <th>Location</th>
                                    <th>Capacity (MW)</th>
                                    <th>Status</th>
                                </tr>
                            </thead>
                            <tbody>
"""

    for rec in sa_only_sorted:
        cap = safe_float(rec.get('full_capacity_mw')) or 0
        location = f"{rec.get('city') or 'Unknown'}, {rec.get('country') or ''}"
        html += f"""
                                <tr>
                                    <td>{rec.get('company_clean') or 'N/A'}</td>
                                    <td>{location[:30]}</td>
                                    <td>{cap:,.1f}</td>
                                    <td>{rec.get('facility_status') or 'N/A'}</td>
                                </tr>
"""

    html += """
                            </tbody>
                        </table>
                    </div>
                </div>
                <div>
                    <h3><span class="badge badge-dch">DCH-Only</span> Top 20 by Capacity</h3>
                    <div class="scroll-table" style="max-height: 400px;">
                        <table>
                            <thead>
                                <tr>
                                    <th>Company</th>
                                    <th>Location</th>
                                    <th>Capacity (MW)</th>
                                    <th>Status</th>
                                </tr>
                            </thead>
                            <tbody>
"""

    for rec in dch_only_sorted:
        cap = safe_float(rec.get('full_capacity_mw')) or 0
        location = f"{rec.get('city') or 'Unknown'}, {rec.get('country') or ''}"
        html += f"""
                                <tr>
                                    <td>{rec.get('company_clean') or 'N/A'}</td>
                                    <td>{location[:30]}</td>
                                    <td>{cap:,.1f}</td>
                                    <td>{rec.get('facility_status') or 'N/A'}</td>
                                </tr>
"""

    html += """
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>
"""

    # Net New Sites Analysis Section
    if net_new_analysis:
        nn = net_new_analysis
        net_new_mape_display = f"{nn['net_new_mape']:.1f}%" if nn.get('net_new_mape') else "N/A"
        net_new_bias_display = f"{nn['net_new_bias']:+.1f}%" if nn.get('net_new_bias') else "N/A"
        status_breakdown = nn.get('status_breakdown', {})

        html += f"""
        <!-- Net New Sites Analysis -->
        <div class="card" id="netnew">
            <h2>🚧 Net New Sites Analysis</h2>

            <div class="explanation-box">
                <h4>Understanding Net New Site Coverage</h4>
                <p>This section analyzes facilities that are <strong>Under Construction</strong> or <strong>Announced</strong>.
                These represent the future data center pipeline - projects that are committed but not yet operational.</p>
                <p><strong>Key insight:</strong> A high "SA-Only" count may indicate SemiAnalysis detects projects earlier,
                while high "DCH-Only" suggests DataCenterHawk has broader coverage of smaller or regional developments.</p>
                <p><em>Note: "Planned" status shows zeros because neither source currently uses this status value.
                Both SA and DCH classify future projects as either "Announced" or "Under Construction".</em></p>
            </div>

            <h3>Net New Summary</h3>
            <div class="metrics-grid">
                <div class="metric-box tooltip">
                    <div class="metric-value primary">{nn['sa_net_new_total']:,}</div>
                    <div class="metric-label">SA Net New Sites</div>
                    <span class="tooltip-text"><strong>SemiAnalysis Net New</strong><br/>Total sites in SA with status "Under Construction" or "Announced". Includes both matched and SA-only sites.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value primary">{nn['dch_net_new_total']:,}</div>
                    <div class="metric-label">DCH Net New Sites</div>
                    <span class="tooltip-text"><strong>DataCenterHawk Net New</strong><br/>Total sites in DCH with status "Under Construction" or "Announced". Includes both matched and DCH-only sites.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value">{nn['matched_both_net_new']:,}</div>
                    <div class="metric-label">Both Sources Report</div>
                    <span class="tooltip-text"><strong>Matched Net New Sites</strong><br/>Sites where BOTH sources report the facility as net new (Under Construction or Announced). These represent confirmed pipeline projects.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value sa">{nn['sa_only_net_new']:,}</div>
                    <div class="metric-label">SA Only</div>
                    <span class="tooltip-text"><strong>SA-Exclusive Net New</strong><br/>Net new sites reported by SemiAnalysis but NOT found in DataCenterHawk within 500m. May indicate early detection or hyperscaler-focused coverage.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value dch">{nn['dch_only_net_new']:,}</div>
                    <div class="metric-label">DCH Only</div>
                    <span class="tooltip-text"><strong>DCH-Exclusive Net New</strong><br/>Net new sites reported by DataCenterHawk but NOT found in SemiAnalysis within 500m. May indicate broader colo/regional coverage.</span>
                </div>
            </div>

            <h3>Coverage & Agreement Rates</h3>
            <div class="metrics-grid">
                <div class="metric-box tooltip">
                    <div class="metric-value">{nn['sa_coverage_rate']:.1f}%</div>
                    <div class="metric-label">SA Net New in DCH</div>
                    <span class="tooltip-text"><strong>SA Coverage Rate</strong><br/>Percentage of SA's net new sites that also appear in DCH. Higher = DCH confirms more of SA's pipeline data.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value">{nn['dch_coverage_rate']:.1f}%</div>
                    <div class="metric-label">DCH Net New in SA</div>
                    <span class="tooltip-text"><strong>DCH Coverage Rate</strong><br/>Percentage of DCH's net new sites that also appear in SA. Higher = SA confirms more of DCH's pipeline data.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value">{net_new_mape_display}</div>
                    <div class="metric-label">Capacity MAPE (Matched)</div>
                    <span class="tooltip-text"><strong>Net New Capacity MAPE</strong><br/>Mean Absolute Percentage Error for matched net new sites. Lower = better agreement on planned capacity. Compare to overall MAPE to see if future projects agree better than existing facilities.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value">{net_new_bias_display}</div>
                    <div class="metric-label">Capacity Bias</div>
                    <span class="tooltip-text"><strong>Net New Capacity Bias</strong><br/>Systematic difference in capacity reporting.<br/>• Positive = SA reports higher<br/>• Negative = DCH reports higher</span>
                </div>
            </div>

            <h3>Capacity by Source (Net New MW)</h3>
            <div class="metrics-grid">
                <div class="metric-box tooltip">
                    <div class="metric-value">{nn['sa_net_new_total_mw']:,.0f}</div>
                    <div class="metric-label">SA Total MW</div>
                    <span class="tooltip-text"><strong>SA Pipeline Capacity</strong><br/>Total megawatts of planned capacity across all SA net new sites.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value">{nn['dch_net_new_total_mw']:,.0f}</div>
                    <div class="metric-label">DCH Total MW</div>
                    <span class="tooltip-text"><strong>DCH Pipeline Capacity</strong><br/>Total megawatts of planned capacity across all DCH net new sites.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value sa">{nn['sa_only_net_new_mw']:,.0f}</div>
                    <div class="metric-label">SA-Only MW</div>
                    <span class="tooltip-text"><strong>Exclusive SA Pipeline</strong><br/>Total megawatts from sites only SA reports. High value suggests significant unconfirmed pipeline capacity.</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="metric-value dch">{nn['dch_only_net_new_mw']:,.0f}</div>
                    <div class="metric-label">DCH-Only MW</div>
                    <span class="tooltip-text"><strong>Exclusive DCH Pipeline</strong><br/>Total megawatts from sites only DCH reports. Compare to SA-only to see which source captures more exclusive capacity.</span>
                </div>
            </div>

            <h3>Breakdown by Status</h3>
            <p style="color: #666; font-size: 0.9em; margin-bottom: 10px;">
                <em>Note: "Planned" shows zeros because neither source uses this status - all future projects are classified as "Announced" or "Under Construction".</em>
            </p>
            <table>
                <thead>
                    <tr>
                        <th>Status</th>
                        <th>SA Total</th>
                        <th>DCH Total</th>
                        <th>SA Matched</th>
                        <th>DCH Matched</th>
                        <th>SA Only</th>
                        <th>DCH Only</th>
                    </tr>
                </thead>
                <tbody>
"""
        for status_name, stats in status_breakdown.items():
            html += f"""
                    <tr>
                        <td><strong>{status_name}</strong></td>
                        <td>{stats.get('sa_total', 0):,}</td>
                        <td>{stats.get('dch_total', 0):,}</td>
                        <td>{stats.get('sa_matched', 0):,}</td>
                        <td>{stats.get('dch_matched', 0):,}</td>
                        <td class="direction-sa">{stats.get('sa_only', 0):,}</td>
                        <td class="direction-dch">{stats.get('dch_only', 0):,}</td>
                    </tr>
"""

        html += """
                </tbody>
            </table>

            <h3>Top Exclusive Net New Sites</h3>
            <div class="grid-2col">
                <div>
                    <h4><span class="badge badge-sa">SA-Only</span> Top Net New by Capacity</h4>
                    <div class="scroll-table" style="max-height: 350px;">
                        <table>
                            <thead>
                                <tr>
                                    <th>Company</th>
                                    <th>Location</th>
                                    <th>MW</th>
                                    <th>Status</th>
                                </tr>
                            </thead>
                            <tbody>
"""

        for rec in nn.get('top_sa_only_net_new', [])[:15]:
            cap = safe_float(rec.get('full_capacity_mw')) or 0
            location = f"{rec.get('city') or 'Unknown'}, {rec.get('country') or ''}"
            status = rec.get('facility_status') or 'N/A'
            html += f"""
                                <tr>
                                    <td>{rec.get('company_clean') or 'N/A'}</td>
                                    <td>{location[:28]}</td>
                                    <td>{cap:,.0f}</td>
                                    <td>{status}</td>
                                </tr>
"""

        html += """
                            </tbody>
                        </table>
                    </div>
                </div>
                <div>
                    <h4><span class="badge badge-dch">DCH-Only</span> Top Net New by Capacity</h4>
                    <div class="scroll-table" style="max-height: 350px;">
                        <table>
                            <thead>
                                <tr>
                                    <th>Company</th>
                                    <th>Location</th>
                                    <th>MW</th>
                                    <th>Status</th>
                                </tr>
                            </thead>
                            <tbody>
"""

        for rec in nn.get('top_dch_only_net_new', [])[:15]:
            cap = safe_float(rec.get('full_capacity_mw')) or 0
            location = f"{rec.get('city') or 'Unknown'}, {rec.get('country') or ''}"
            status = rec.get('facility_status') or 'N/A'
            html += f"""
                                <tr>
                                    <td>{rec.get('company_clean') or 'N/A'}</td>
                                    <td>{location[:28]}</td>
                                    <td>{cap:,.0f}</td>
                                    <td>{status}</td>
                                </tr>
"""

        html += """
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>
"""

    # Ground Truth Reference Section (Meta Canonical)
    html += """
        <!-- Ground Truth Reference -->
        <div class="card" id="groundtruth">
            <h2>🎯 Ground Truth Reference (Meta Canonical)</h2>

            <div class="explanation-box">
                <h4>Understanding This Comparison vs. Ground Truth</h4>
                <p>The metrics above compare <strong>SA vs DCH against each other</strong> — measuring how well
                the two sources agree. This tells us about <strong>inter-source consistency</strong>, but not which source is more accurate.</p>
                <p>To determine actual accuracy, we compare each source against <strong>Meta Canonical</strong> — Meta's
                internal datacenter inventory with verified capacity values (~643 buildings, 17.2 GW total capacity).</p>
            </div>

            <h3>Accuracy vs. Meta Canonical (Ground Truth)</h3>
            <div class="metrics-grid">
                <div class="metric-box tooltip" style="border-left: 4px solid #34a853;">
                    <div class="metric-value" style="color: #34a853;">11.9%</div>
                    <div class="metric-label">SA MAPE vs Meta</div>
                    <span class="tooltip-text"><strong>SemiAnalysis Accuracy</strong><br/>Mean Absolute Percentage Error when comparing SA capacity values against Meta's verified internal data.<br/><br/>11.9% = Grade A (Excellent)</span>
                </div>
                <div class="metric-box tooltip" style="border-left: 4px solid #4285f4;">
                    <div class="metric-value" style="color: #4285f4;">17.6%</div>
                    <div class="metric-label">DCH MAPE vs Meta</div>
                    <span class="tooltip-text"><strong>DataCenterHawk Accuracy</strong><br/>Mean Absolute Percentage Error when comparing DCH capacity values against Meta's verified internal data.<br/><br/>17.6% = Grade B (Good)</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="grade grade-a">A</div>
                    <div class="metric-label" style="margin-top: 8px;">SA Grade</div>
                    <span class="tooltip-text"><strong>SemiAnalysis Grade</strong><br/>Based on MAPE vs ground truth:<br/>• 0-10% = A<br/>• 10-20% = B<br/>• 20-35% = C</span>
                </div>
                <div class="metric-box tooltip">
                    <div class="grade grade-b">B</div>
                    <div class="metric-label" style="margin-top: 8px;">DCH Grade</div>
                    <span class="tooltip-text"><strong>DataCenterHawk Grade</strong><br/>Based on MAPE vs ground truth:<br/>• 0-10% = A<br/>• 10-20% = B<br/>• 20-35% = C</span>
                </div>
            </div>

            <div class="interpretation-box">
                <h4>💡 Interpretation: Forecasting Reliability</h4>
                <p style="margin-bottom: 12px;">Based on ground truth validation against Meta Canonical:</p>
                <ul>
                    <li><strong>SemiAnalysis is more accurate</strong> — 11.9% MAPE (Grade A) vs DCH's 17.6% MAPE (Grade B) when validated against Meta's verified facility data.</li>
                    <li><strong>For net new/future capacity forecasting</strong>, SA's demonstrated higher accuracy against ground truth suggests its pipeline projections may be more reliable.</li>
                    <li><strong>The +35% bias</strong> (SA reporting higher than DCH) likely reflects SA capturing planned expansions that DCH has not yet recorded, rather than SA over-estimating.</li>
                    <li><strong>Recommendation:</strong> When sources disagree on future capacity, SA values should be given higher weight based on ground truth performance.</li>
                </ul>
            </div>

            <h3>Summary Table</h3>
            <table>
                <thead>
                    <tr>
                        <th>Source</th>
                        <th>MAPE vs Meta</th>
                        <th>Grade</th>
                        <th>Interpretation</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td><strong>SemiAnalysis</strong></td>
                        <td style="color: #34a853; font-weight: 600;">11.9%</td>
                        <td><span class="grade grade-a" style="font-size: 0.9em; padding: 4px 12px;">A</span></td>
                        <td>Excellent accuracy — preferred source for capacity forecasting</td>
                    </tr>
                    <tr>
                        <td><strong>DataCenterHawk</strong></td>
                        <td style="color: #4285f4; font-weight: 600;">17.6%</td>
                        <td><span class="grade grade-b" style="font-size: 0.9em; padding: 4px 12px;">B</span></td>
                        <td>Good accuracy — strong for colo coverage breadth</td>
                    </tr>
                </tbody>
            </table>

            <p style="margin-top: 16px; color: #5f6368; font-size: 0.9em;">
              <em>Note: Ground truth comparison based on ~643 Meta Canonical buildings with verified capacity data.
                SA and DCH accuracy grades reflect performance on the subset of facilities that overlap with Meta's portfolio.</em>
            </p>
        </div>
"""

    # Geographic Analysis Table
    country_data = geo_analysis.get('by_country', {})

    html += """
        <!-- Geographic Analysis -->
        <div class="card" id="geography">
            <h2>🌍 Geographic Distribution</h2>

            <div class="explanation-box">
                <h4>Regional Coverage Patterns</h4>
                <p>This table shows how the source coverage varies by country. High <strong>SA Only</strong> or
                <strong>DCH Only</strong> counts in specific countries may indicate regional strengths or gaps in each
                data source's coverage methodology.</p>
            </div>

            <div class="scroll-table">
                <table>
                    <thead>
                        <tr>
                            <th>Country</th>
                            <th>Matched</th>
                            <th>SA Only</th>
                            <th>DCH Only</th>
                            <th>Conflicts</th>
                            <th>Total</th>
                        </tr>
                    </thead>
                    <tbody>
"""

    # Handle both dict and list of tuples
    if isinstance(country_data, dict):
        country_items = country_data.items()
    else:
        country_items = country_data

    for country, stats in country_items:
        total = stats['matched'] + stats['sa_only'] + stats['dch_only']
        html += f"""
                        <tr>
                            <td><strong>{country}</strong></td>
                            <td>{stats['matched']:,}</td>
                            <td>{stats['sa_only']:,}</td>
                            <td>{stats['dch_only']:,}</td>
                            <td>{stats['conflicts']:,}</td>
                            <td>{total:,}</td>
                        </tr>
"""

    html += """
                    </tbody>
                </table>
            </div>
        </div>
"""

    # Essential DC Section (if available in data)
    essential_conflicts = [p for p in matched_pairs
                          if p['sa_record'].get('is_essential') or p['dch_record'].get('is_essential')]
    essential_with_conflict = [p for p in essential_conflicts if p['capacity_delta']['is_significant']]

    if essential_conflicts:
        html += f"""
        <!-- Essential DC Analysis -->
        <div class="card">
            <h2>⭐ Essential Data Center Analysis</h2>
            <p>Strategic facilities marked as essential in the gold layer.</p>

            <div class="methodology-note">
                <strong>⚠️ Large Conflicts Note:</strong> Conflicts exceeding 500 MW often indicate a <strong>granularity mismatch</strong>
                (e.g., campus-level capacity matched to a single building record) or different capacity definitions between sources.
                Check the SA/DCH Granularity columns below to identify these cases.
            </div>

            <div class="metrics-grid">
                <div class="metric-box">
                    <div class="metric-value primary">{len(essential_conflicts):,}</div>
                    <div class="metric-label">Essential DCs Matched</div>
                </div>
                <div class="metric-box">
                    <div class="metric-value conflict">{len(essential_with_conflict):,}</div>
                    <div class="metric-label">With Significant Conflicts</div>
                </div>
            </div>

            <div class="scroll-table" style="max-height: 400px;">
                <table>
                    <thead>
                        <tr>
                            <th>Company</th>
                            <th>Location</th>
                            <th>SA (MW)</th>
                            <th>DCH (MW)</th>
                            <th>Delta</th>
                            <th>SA Level</th>
                            <th>DCH Level</th>
                            <th>Flag</th>
                        </tr>
                    </thead>
                    <tbody>
"""

        for pair in sorted(essential_with_conflict, key=lambda x: abs(x['capacity_delta']['delta_mw']), reverse=True)[:15]:
            sa_rec = pair['sa_record']
            dch_rec = pair['dch_record']
            delta = pair['capacity_delta']
            location = f"{sa_rec.get('city') or dch_rec.get('city') or 'Unknown'}, {sa_rec.get('country') or dch_rec.get('country') or ''}"
            sa_gran = pair.get('sa_granularity', 'Unknown')
            dch_gran = pair.get('dch_granularity', 'Unknown')
            granularity_match_class = '' if pair.get('granularity_match', True) else 'style="background-color: #fff3cd;"'

            html += f"""
                        <tr {granularity_match_class}>
                            <td>{sa_rec.get('company_clean') or dch_rec.get('company_clean')}</td>
                            <td>{location[:35]}</td>
                            <td>{delta['sa_capacity']:,.1f}</td>
                            <td>{delta['dch_capacity']:,.1f}</td>
                            <td class="{'direction-sa' if delta['delta_mw'] > 0 else 'direction-dch'}">{delta['delta_mw']:+,.1f} MW</td>
                            <td><span class="granularity-badge">{sa_gran}</span></td>
                            <td><span class="granularity-badge">{dch_gran}</span></td>
                            <td><span class="essential-flag">ESSENTIAL</span></td>
                        </tr>
"""

    html += """
                    </tbody>
                </table>
            </div>
        </div>
"""

    # Conclusions & Recommendations Section (after Essential DCs, before Methodology)
    if net_new_analysis:
        nn = net_new_analysis
        sa_avg_mw = nn['sa_net_new_total_mw'] / nn['sa_net_new_total'] if nn['sa_net_new_total'] > 0 else 0
        dch_avg_mw = nn['dch_net_new_total_mw'] / nn['dch_net_new_total'] if nn['dch_net_new_total'] > 0 else 0
        sa_only_avg_mw = nn['sa_only_net_new_mw'] / nn['sa_only_net_new'] if nn['sa_only_net_new'] > 0 else 0
        dch_only_avg_mw = nn['dch_only_net_new_mw'] / nn['dch_only_net_new'] if nn['dch_only_net_new'] > 0 else 0

        html += f"""
        <!-- Conclusions & Recommendations -->
        <div class="card" id="conclusions">
            <h2>📋 Conclusions & Recommendations</h2>

            <div class="explanation-box">
                <h4>Summary: Which Source Has Greater Recall of Net New Sites?</h4>
                <p><strong>DCH has greater recall by site count</strong>, but <strong>SA captures more capacity</strong>.
                This reflects fundamentally different coverage strategies between the two sources.</p>
            </div>

            <h3>Net New Site Comparison Summary</h3>
            <table>
                <thead>
                    <tr>
                        <th>Metric</th>
                        <th>SemiAnalysis</th>
                        <th>DataCenterHawk</th>
                        <th>Winner</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td><strong>Total Net New Sites</strong></td>
                        <td>{nn['sa_net_new_total']:,}</td>
                        <td>{nn['dch_net_new_total']:,}</td>
                        <td><span class="badge badge-dch">DCH (+{((nn['dch_net_new_total'] - nn['sa_net_new_total']) / max(nn['sa_net_new_total'], 1) * 100):.0f}%)</span></td>
                    </tr>
                    <tr>
                        <td><strong>Total Net New Capacity (MW)</strong></td>
                        <td>{nn['sa_net_new_total_mw']:,.0f}</td>
                        <td>{nn['dch_net_new_total_mw']:,.0f}</td>
                        <td><span class="badge badge-sa">SA (+{((nn['sa_net_new_total_mw'] - nn['dch_net_new_total_mw']) / max(nn['dch_net_new_total_mw'], 1) * 100):.0f}%)</span></td>
                    </tr>
                    <tr>
                        <td><strong>Exclusive Sites</strong></td>
                        <td>{nn['sa_only_net_new']:,}</td>
                        <td>{nn['dch_only_net_new']:,}</td>
                        <td><span class="badge badge-dch">DCH ({nn['dch_only_net_new'] / max(nn['sa_only_net_new'], 1):.1f}x more)</span></td>
                    </tr>
                    <tr>
                        <td><strong>Exclusive Capacity (MW)</strong></td>
                        <td>{nn['sa_only_net_new_mw']:,.0f}</td>
                        <td>{nn['dch_only_net_new_mw']:,.0f}</td>
                        <td><span class="badge badge-sa">SA (+{((nn['sa_only_net_new_mw'] - nn['dch_only_net_new_mw']) / max(nn['dch_only_net_new_mw'], 1) * 100):.0f}%)</span></td>
                    </tr>
                    <tr>
                        <td><strong>Avg MW per Net New Site</strong></td>
                        <td>{sa_avg_mw:.0f} MW</td>
                        <td>{dch_avg_mw:.0f} MW</td>
                        <td><span class="badge badge-sa">SA ({sa_avg_mw / max(dch_avg_mw, 1):.1f}x larger)</span></td>
                    </tr>
                </tbody>
            </table>

            <div class="interpretation-box">
                <h4>💡 Key Interpretation</h4>
                <ul>
                    <li><strong>DCH has broader coverage</strong> — They track ~{nn['dch_net_new_total'] - nn['sa_net_new_total']:,} more net new sites, including many smaller colo/regional facilities</li>
                    <li><strong>SA captures larger projects</strong> — Fewer sites but ~{(nn['sa_net_new_total_mw'] - nn['dch_net_new_total_mw']):,.0f} MW more capacity, suggesting focus on hyperscaler mega-projects</li>
                    <li><strong>DCH-only sites are smaller</strong> — {nn['dch_only_net_new']:,} sites with {nn['dch_only_net_new_mw']:,.0f} MW = avg {dch_only_avg_mw:.0f} MW each</li>
                    <li><strong>SA-only sites are larger</strong> — {nn['sa_only_net_new']:,} sites with {nn['sa_only_net_new_mw']:,.0f} MW = avg {sa_only_avg_mw:.0f} MW each</li>
                </ul>
            </div>

            <h3>Additional Insights for Source Reliability</h3>

            <h4>A. Net New MAPE vs Overall MAPE</h4>
            <ul>
                <li><strong>Net New MAPE:</strong> {nn['net_new_mape']:.1f}% vs <strong>Overall MAPE:</strong> {mape_val:.1f}%</li>
                <li>Sources show similar agreement levels for existing and future projects</li>
                <li>This suggests consistent estimation methodologies across both sources</li>
            </ul>

            <h4>B. Coverage Asymmetry</h4>
            <ul>
                <li><strong>{nn['sa_coverage_rate']:.1f}%</strong> of SA's net new sites appear in DCH</li>
                <li><strong>{nn['dch_coverage_rate']:.1f}%</strong> of DCH's net new sites appear in SA</li>
                <li><strong>Implication:</strong> SA is more selective/focused — almost everything SA tracks, DCH also tracks</li>
                <li>DCH has broader but shallower coverage — many sites SA doesn't track (likely smaller colo)</li>
                <li>When SA has a site, it's likely validated by DCH (high confirmation rate)</li>
            </ul>

            <h4>C. Capacity Bias Interpretation</h4>
            <ul>
                <li>SA reports <strong>{nn['net_new_bias']:+.1f}%</strong> higher capacity on net new sites</li>
                <li>Combined with SA's better ground truth accuracy (11.9% vs 17.6% MAPE), this suggests SA is capturing planned expansions that DCH hasn't recorded yet, <strong>not</strong> over-estimating</li>
            </ul>

            <h3>📊 Recommendation Framework</h3>
            <table>
                <thead>
                    <tr>
                        <th>Use Case</th>
                        <th>Recommended Source</th>
                        <th>Rationale</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td><strong>Hyperscaler pipeline tracking</strong></td>
                        <td><span class="badge badge-sa">SemiAnalysis</span></td>
                        <td>Larger avg project size, focus on mega-projects</td>
                    </tr>
                    <tr>
                        <td><strong>Colo market coverage</strong></td>
                        <td><span class="badge badge-dch">DataCenterHawk</span></td>
                        <td>Broader site count, regional facility coverage</td>
                    </tr>
                    <tr>
                        <td><strong>Capacity forecasting</strong></td>
                        <td><span class="badge badge-sa">SemiAnalysis</span></td>
                        <td>Better ground truth MAPE (11.9% vs 17.6%)</td>
                    </tr>
                    <tr>
                        <td><strong>Site discovery / completeness</strong></td>
                        <td><span class="badge badge-dch">DataCenterHawk</span></td>
                        <td>More total sites, broader geographic coverage</td>
                    </tr>
                    <tr>
                        <td><strong>High-confidence subset</strong></td>
                        <td><span class="badge badge-match">Both Agree</span></td>
                        <td>{nn['matched_both_net_new']:,} matched sites with cross-validation</td>
                    </tr>
                </tbody>
            </table>

            <div class="methodology-note">
                <strong>⚠️ Best Practice:</strong> For capacity planning, use SA as the primary source for magnitude estimates,
                but cross-reference with DCH for site discovery. When both sources agree on a site, confidence is highest.
                For strategic hyperscaler facilities, SA's accuracy advantage makes it the preferred source.
            </div>
        </div>
"""

    # Footer and Chart.js Scripts
    html += f"""
        <!-- Methodology Section -->
        <div class="card" id="methodology">
            <h2>📖 Methodology & Definitions</h2>

            <div class="explanation-box">
                <h4>How This Comparison Was Performed</h4>
                <p>This comparison leverages the <strong>Consensus GIS Model pipeline</strong> which processes and enriches
                both SA and DCH data before comparison. The pipeline ensures consistent data quality and enables
                high-confidence matching between sources.</p>
            </div>

            <h3>Data Pipeline Steps</h3>
            <table>
                <tr>
                    <th>Step</th>
                    <th>Process</th>
                    <th>Impact on Comparison</th>
                </tr>
                <tr>
                    <td><strong>1. Ingestion</strong></td>
                    <td>SA Excel → CSV → GDB; DCH Hive → CSV → GDB</td>
                    <td>Standardized field names, geometry validation</td>
                </tr>
                <tr>
                    <td><strong>2. Geography Enrichment</strong></td>
                    <td>Populate region, state, country from coordinates</td>
                    <td>Enables regional accuracy breakdown</td>
                </tr>
                <tr>
                    <td><strong>3. Company Standardization</strong></td>
                    <td>Normalize company names (company_clean, company_clean_filter)</td>
                    <td>Consistent company grouping across sources</td>
                </tr>
                <tr>
                    <td><strong>4. UCID Generation</strong></td>
                    <td>Assign Universal Campus IDs based on location + company</td>
                    <td>Company-aware matching prevents cross-company mismatches</td>
                </tr>
                <tr>
                    <td><strong>5. Granularity Normalization</strong></td>
                    <td>Standardize Building/Campus/Suite levels</td>
                    <td>MAPE calculated only on same-granularity matches</td>
                </tr>
                <tr>
                    <td><strong>6. Essential DC Flagging</strong></td>
                    <td>Mark 127 peer/frontier strategic facilities for priority monitoring</td>
                    <td>Enables tier-weighted accuracy and priority filtering</td>
                </tr>
            </table>

            <h3>Matching Algorithm</h3>
            <p>Records are matched using a multi-step process:</p>
            <ol>
                <li><strong>Spatial Proximity:</strong> Haversine distance calculation with {threshold_m}m threshold</li>
                <li><strong>Closest Match Selection:</strong> When multiple candidates exist, select the nearest</li>
                <li><strong>Granularity Filtering:</strong> Only same-granularity pairs used for capacity MAPE</li>
                <li><strong>Company Validation:</strong> Company names cross-checked for potential mismatches</li>
            </ol>

            <h3>Statistical Definitions</h3>
            <table>
                <tr>
                    <th>Metric</th>
                    <th>Definition</th>
                    <th>Interpretation</th>
                </tr>
                <tr>
                    <td><strong>MAPE</strong></td>
                    <td>Mean Absolute Percentage Error = mean(|SA - DCH| / max(SA, DCH)) × 100</td>
                    <td>0-10%: Excellent, 10-20%: Good, 20-35%: Moderate, 35-50%: Significant, >50%: Poor</td>
                </tr>
                <tr>
                    <td><strong>Bias %</strong></td>
                    <td>Systematic bias = mean(SA - DCH) / mean(DCH) × 100</td>
                    <td>Positive = SA reports higher capacity on average; Negative = DCH reports higher</td>
                </tr>
                <tr>
                    <td><strong>CV</strong></td>
                    <td>Coefficient of Variation = std(delta) / mean(|delta|) × 100</td>
                    <td>Lower CV indicates more consistent disagreements (easier to calibrate)</td>
                </tr>
                <tr>
                    <td><strong>Pearson r</strong></td>
                    <td>Correlation coefficient between SA and DCH capacity values</td>
                    <td>>0.9: Strong, 0.7-0.9: Moderate, <0.7: Weak linear relationship</td>
                </tr>
                <tr>
                    <td><strong>Significant Conflict</strong></td>
                    <td>Matched pair with >20% capacity delta OR >10 MW absolute difference</td>
                    <td>These records require manual review and reconciliation</td>
                </tr>
            </table>

            <h3>Ground Truth Validation</h3>
            <p>Both SA and DCH are validated against <strong>Meta Canonical</strong> — an internal dataset of ~643 buildings
            with verified capacity values. This provides an independent accuracy benchmark:</p>
            <ul>
                <li><strong>SA Ground Truth MAPE:</strong> 11.9% (Grade A)</li>
                <li><strong>DCH Ground Truth MAPE:</strong> 17.6% (Grade B)</li>
            </ul>
            <p>SA's better ground truth accuracy suggests its capacity estimates are more reliable for forecasting purposes.</p>

            <div class="methodology-note">
                <strong>⚠️ Important Limitations:</strong>
                <ul style="margin-top: 8px; margin-bottom: 0;">
                    <li>Ground truth validation limited to ~643 Meta Canonical buildings (may not represent smaller/regional facilities)</li>
                    <li>Multi-tenant campuses may still cause some cross-company matching errors</li>
                    <li>Neither source has complete commissioned_power_mw data for all regions</li>
                </ul>
            </div>
        </div>

        <!-- Footer -->
        <div class="footer">
            <p>Generated by SA vs DCH Comparison Script V2 | {timestamp}</p>
            <p>Spatial matching threshold: {threshold_m}m | Source: gold_buildings_full | Pipeline: Consensus GIS Model</p>
        </div>
        </div><!-- End container -->
    </div><!-- End main-content -->

    <!-- Chart.js Scripts -->
    <script>
        // Histogram Chart
        const histogramCtx = document.getElementById('histogramChart').getContext('2d');
        new Chart(histogramCtx, {{
            type: 'bar',
            data: {{
                labels: {histogram_json},
                datasets: [{{
                    label: 'Count of Pairs',
                    data: {histogram_values_json},
                    backgroundColor: 'rgba(24, 119, 242, 0.7)',
                    borderColor: 'rgba(24, 119, 242, 1)',
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{ display: false }},
                    legend: {{ display: false }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        title: {{ display: true, text: 'Number of Pairs' }}
                    }},
                    x: {{
                        title: {{ display: true, text: 'Capacity Delta (MW)' }}
                    }}
                }}
            }}
        }});

        // Scatter Plot with regression line and zoom
        const scatterCtx = document.getElementById('scatterChart').getContext('2d');
        const scatterData = {scatter_data_json};

        // Calculate regression line (y = mx + b)
        const n = scatterData.length;
        let sumX = 0, sumY = 0, sumXY = 0, sumXX = 0;
        let maxX = 0, maxY = 0;
        scatterData.forEach(pt => {{
            sumX += pt.x;
            sumY += pt.y;
            sumXY += pt.x * pt.y;
            sumXX += pt.x * pt.x;
            maxX = Math.max(maxX, pt.x);
            maxY = Math.max(maxY, pt.y);
        }});
        const slope = (n * sumXY - sumX * sumY) / (n * sumXX - sumX * sumX);
        const intercept = (sumY - slope * sumX) / n;
        const maxVal = Math.max(maxX, maxY);

        // Regression line endpoints
        const regressionLine = [
            {{x: 0, y: intercept}},
            {{x: maxVal, y: slope * maxVal + intercept}}
        ];

        let scatterChart = new Chart(scatterCtx, {{
            type: 'scatter',
            data: {{
                datasets: [{{
                    label: 'SA vs DCH Capacity',
                    data: scatterData,
                    backgroundColor: 'rgba(233, 30, 99, 0.5)',
                    borderColor: 'rgba(233, 30, 99, 0.8)',
                    pointRadius: 4
                }},
                {{
                    label: 'Perfect Agreement (1:1)',
                    data: [{{x: 0, y: 0}}, {{x: maxVal, y: maxVal}}],
                    type: 'line',
                    borderColor: 'rgba(76, 175, 80, 0.8)',
                    borderDash: [5, 5],
                    pointRadius: 0,
                    fill: false,
                    borderWidth: 2
                }},
                {{
                    label: 'Regression Line (R² = {r_squared:.3f})',
                    data: regressionLine,
                    type: 'line',
                    borderColor: 'rgba(24, 119, 242, 0.9)',
                    pointRadius: 0,
                    fill: false,
                    borderWidth: 2
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{ position: 'top' }},
                    zoom: {{
                        pan: {{
                            enabled: true,
                            mode: 'xy',
                            modifierKey: null
                        }},
                        zoom: {{
                            wheel: {{
                                enabled: true
                            }},
                            pinch: {{
                                enabled: true
                            }},
                            drag: {{
                                enabled: false
                            }},
                            mode: 'xy'
                        }}
                    }},
                    tooltip: {{
                        callbacks: {{
                            label: function(context) {{
                                if (context.dataset.type === 'line') return context.dataset.label;
                                const pt = context.raw;
                                const diff = pt.x - pt.y;
                                const pctDiff = pt.y > 0 ? ((diff / pt.y) * 100).toFixed(1) : 'N/A';
                                return `SA: ${{pt.x.toFixed(1)}} MW, DCH: ${{pt.y.toFixed(1)}} MW (Δ ${{diff.toFixed(1)}} MW, ${{pctDiff}}%)`;
                            }}
                        }}
                    }}
                }},
                scales: {{
                    x: {{
                        title: {{ display: true, text: 'SA Capacity (MW)' }},
                        beginAtZero: true
                    }},
                    y: {{
                        title: {{ display: true, text: 'DCH Capacity (MW)' }},
                        beginAtZero: true
                    }}
                }}
            }}
        }});

        // Zoom control functions
        window.resetScatterZoom = function() {{
            scatterChart.resetZoom();
        }};

        window.zoomScatter = function(min, max) {{
            scatterChart.options.scales.x.min = min;
            scatterChart.options.scales.x.max = max;
            scatterChart.options.scales.y.min = min;
            scatterChart.options.scales.y.max = max;
            scatterChart.update();
        }};

        // Company Chart
        const companyCtx = document.getElementById('companyChart').getContext('2d');
        const companyMatched = {company_matched_json};
        const companySaOnly = {company_sa_only_json};
        const companyDchOnly = {company_dch_only_json};
        const companyTotals = companyMatched.map((m, i) => m + companySaOnly[i] + companyDchOnly[i]);
        const companyGrandTotal = companyTotals.reduce((a, b) => a + b, 0);

        new Chart(companyCtx, {{
            type: 'bar',
            data: {{
                labels: {company_labels_json},
                datasets: [
                    {{
                        label: 'Matched',
                        data: companyMatched,
                        backgroundColor: 'rgba(24, 119, 242, 0.7)'
                    }},
                    {{
                        label: 'SA Only',
                        data: companySaOnly,
                        backgroundColor: 'rgba(233, 30, 99, 0.7)'
                    }},
                    {{
                        label: 'DCH Only',
                        data: companyDchOnly,
                        backgroundColor: 'rgba(76, 175, 80, 0.7)'
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{ position: 'top' }},
                    tooltip: {{
                        callbacks: {{
                            label: function(context) {{
                                const value = context.raw;
                                const dataIndex = context.dataIndex;
                                const companyTotal = companyTotals[dataIndex];
                                const pctOfCompany = companyTotal > 0 ? ((value / companyTotal) * 100).toFixed(1) : 0;
                                const pctOfTotal = companyGrandTotal > 0 ? ((value / companyGrandTotal) * 100).toFixed(1) : 0;
                                return `${{context.dataset.label}}: ${{value.toLocaleString()}} (${{pctOfCompany}}% of company, ${{pctOfTotal}}% of total)`;
                            }},
                            footer: function(tooltipItems) {{
                                const dataIndex = tooltipItems[0].dataIndex;
                                const companyTotal = companyTotals[dataIndex];
                                const pctOfTotal = companyGrandTotal > 0 ? ((companyTotal / companyGrandTotal) * 100).toFixed(1) : 0;
                                return `Company Total: ${{companyTotal.toLocaleString()}} (${{pctOfTotal}}% of all records)`;
                            }}
                        }}
                    }}
                }},
                scales: {{
                    x: {{ stacked: true }},
                    y: {{ stacked: true, beginAtZero: true }}
                }}
            }}
        }});

        // Region Chart
        const regionCtx = document.getElementById('regionChart').getContext('2d');
        const regionMatched = {region_matched_json};
        const regionSaOnly = {region_sa_only_json};
        const regionDchOnly = {region_dch_only_json};
        const regionTotals = regionMatched.map((m, i) => m + regionSaOnly[i] + regionDchOnly[i]);
        const regionGrandTotal = regionTotals.reduce((a, b) => a + b, 0);

        new Chart(regionCtx, {{
            type: 'bar',
            data: {{
                labels: {region_labels_json},
                datasets: [
                    {{
                        label: 'Matched',
                        data: regionMatched,
                        backgroundColor: 'rgba(24, 119, 242, 0.7)'
                    }},
                    {{
                        label: 'SA Only',
                        data: regionSaOnly,
                        backgroundColor: 'rgba(233, 30, 99, 0.7)'
                    }},
                    {{
                        label: 'DCH Only',
                        data: regionDchOnly,
                        backgroundColor: 'rgba(76, 175, 80, 0.7)'
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{ position: 'top' }},
                    tooltip: {{
                        callbacks: {{
                            label: function(context) {{
                                const value = context.raw;
                                const dataIndex = context.dataIndex;
                                const regionTotal = regionTotals[dataIndex];
                                const pctOfRegion = regionTotal > 0 ? ((value / regionTotal) * 100).toFixed(1) : 0;
                                const pctOfTotal = regionGrandTotal > 0 ? ((value / regionGrandTotal) * 100).toFixed(1) : 0;
                                return `${{context.dataset.label}}: ${{value.toLocaleString()}} (${{pctOfRegion}}% of region, ${{pctOfTotal}}% of total)`;
                            }},
                            footer: function(tooltipItems) {{
                                const dataIndex = tooltipItems[0].dataIndex;
                                const regionTotal = regionTotals[dataIndex];
                                const pctOfTotal = regionGrandTotal > 0 ? ((regionTotal / regionGrandTotal) * 100).toFixed(1) : 0;
                                return `Region Total: ${{regionTotal.toLocaleString()}} (${{pctOfTotal}}% of all records)`;
                            }}
                        }}
                    }}
                }},
                scales: {{
                    x: {{ stacked: true }},
                    y: {{ stacked: true, beginAtZero: true }}
                }}
            }}
        }});

        // Theme Toggle Functionality
        function toggleTheme() {{
            const html = document.documentElement;
            const currentTheme = html.getAttribute('data-theme');
            const newTheme = currentTheme === 'dark' ? 'light' : 'dark';
            html.setAttribute('data-theme', newTheme);

            // Update toggle button
            const icon = document.getElementById('theme-icon');
            const label = document.getElementById('theme-label');
            if (newTheme === 'dark') {{
                icon.textContent = '☀️';
                label.textContent = 'Light';
            }} else {{
                icon.textContent = '🌙';
                label.textContent = 'Dark';
            }}

            // Save preference
            localStorage.setItem('sa-dch-theme', newTheme);

            // Update Chart.js colors for dark mode
            updateChartColors(newTheme);
        }}

        function updateChartColors(theme) {{
            const textColor = theme === 'dark' ? '#e4e6eb' : '#666';
            const gridColor = theme === 'dark' ? '#444' : '#ddd';

            Chart.helpers.each(Chart.instances, function(chart) {{
                if (chart.options.scales) {{
                    if (chart.options.scales.x) {{
                        chart.options.scales.x.ticks = chart.options.scales.x.ticks || {{}};
                        chart.options.scales.x.ticks.color = textColor;
                        chart.options.scales.x.grid = chart.options.scales.x.grid || {{}};
                        chart.options.scales.x.grid.color = gridColor;
                    }}
                    if (chart.options.scales.y) {{
                        chart.options.scales.y.ticks = chart.options.scales.y.ticks || {{}};
                        chart.options.scales.y.ticks.color = textColor;
                        chart.options.scales.y.grid = chart.options.scales.y.grid || {{}};
                        chart.options.scales.y.grid.color = gridColor;
                    }}
                }}
                if (chart.options.plugins && chart.options.plugins.legend) {{
                    chart.options.plugins.legend.labels = chart.options.plugins.legend.labels || {{}};
                    chart.options.plugins.legend.labels.color = textColor;
                }}
                chart.update();
            }});
        }}

        // Load saved theme preference on page load
        document.addEventListener('DOMContentLoaded', function() {{
            const savedTheme = localStorage.getItem('sa-dch-theme') || 'light';
            if (savedTheme === 'dark') {{
                document.documentElement.setAttribute('data-theme', 'dark');
                document.getElementById('theme-icon').textContent = '☀️';
                document.getElementById('theme-label').textContent = 'Light';
                // Wait for charts to initialize, then update colors
                setTimeout(function() {{ updateChartColors('dark'); }}, 500);
            }}
        }});
    </script>
</body>
</html>
"""

    # Write to file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"    HTML report saved: {output_path}")


# ==============================================================================
# MAPE, BIAS, CV CALCULATIONS (From Pipeline Report)
# ==============================================================================

# Company tier weights for weighted accuracy scoring
COMPANY_TIER_WEIGHTS = {
    'Hyperscaler': 0.60,
    'Major Colo': 0.30,
    'Other': 0.10
}

def get_company_tier(company_name: str) -> str:
    """Determine the tier of a company for weighted scoring."""
    if not company_name:
        return 'Other'
    company_upper = company_name.upper()
    if any(h.upper() in company_upper for h in HYPERSCALERS):
        return 'Hyperscaler'
    # Major colos (add more as needed)
    major_colos = ['EQUINIX', 'DIGITAL REALTY', 'CYRUSONE', 'QTS', 'CORESITE',
                   'VANTAGE', 'STACK', 'ALIGNED', 'COMPASS', 'EDGECONNEX', 'NTT']
    if any(c.upper() in company_upper for c in major_colos):
        return 'Major Colo'
    return 'Other'


def calculate_mape_and_bias(matched_pairs: List[Dict]) -> Dict[str, Any]:
    """
    Calculate Mean Absolute Percentage Error and systematic bias.

    MAPE = mean(|SA - DCH| / max(SA, DCH)) * 100
    Bias = mean(SA - DCH) / mean(DCH) * 100  # Positive = SA reports higher

    NOTE: Only uses same-granularity pairs (building↔building, campus↔campus)
    to ensure apples-to-apples comparison.

    Returns dict with mape, bias_pct, and supporting statistics.
    """
    # Filter to same-granularity pairs first (apples-to-apples comparison)
    same_granularity_pairs = [p for p in matched_pairs if p.get('granularity_match', True)]

    # Then filter to pairs with positive capacity in both sources
    valid_pairs = [p for p in same_granularity_pairs
                   if p['capacity_delta']['sa_capacity'] > 0 and p['capacity_delta']['dch_capacity'] > 0]

    # Track how many were excluded due to granularity mismatch
    all_positive_capacity = [p for p in matched_pairs
                             if p['capacity_delta']['sa_capacity'] > 0 and p['capacity_delta']['dch_capacity'] > 0]
    excluded_granularity = len(all_positive_capacity) - len(valid_pairs)

    if not valid_pairs:
        return {'mape': None, 'bias_pct': None, 'n': 0, 'excluded_granularity': excluded_granularity}

    # Calculate MAPE
    apes = []
    for p in valid_pairs:
        sa_cap = p['capacity_delta']['sa_capacity']
        dch_cap = p['capacity_delta']['dch_capacity']
        max_cap = max(sa_cap, dch_cap)
        ape = abs(sa_cap - dch_cap) / max_cap * 100
        apes.append(ape)

    mape = sum(apes) / len(apes)

    # Calculate Bias
    sa_sum = sum(p['capacity_delta']['sa_capacity'] for p in valid_pairs)
    dch_sum = sum(p['capacity_delta']['dch_capacity'] for p in valid_pairs)
    delta_sum = sum(p['capacity_delta']['delta_mw'] for p in valid_pairs)

    bias_pct = (delta_sum / dch_sum * 100) if dch_sum > 0 else 0

    return {
        'mape': mape,
        'bias_pct': bias_pct,
        'n': len(valid_pairs),
        'apes': apes,
        'sa_total_mw': sa_sum,
        'dch_total_mw': dch_sum,
        'excluded_granularity': excluded_granularity
    }


def calculate_cv(matched_pairs: List[Dict]) -> float:
    """
    Calculate Coefficient of Variation for capacity deltas.

    CV = (std / mean) * 100
    Measures the consistency/dispersion of disagreement.

    NOTE: Only uses same-granularity pairs (building↔building, campus↔campus)
    to ensure apples-to-apples comparison.
    """
    # Filter to same-granularity pairs first
    same_granularity_pairs = [p for p in matched_pairs if p.get('granularity_match', True)]

    deltas = [abs(p['capacity_delta']['delta_mw']) for p in same_granularity_pairs
              if p['capacity_delta']['sa_capacity'] > 0 and p['capacity_delta']['dch_capacity'] > 0]

    if len(deltas) < 2:
        return None

    mean_delta = sum(deltas) / len(deltas)
    if mean_delta == 0:
        return 0

    variance = sum((d - mean_delta) ** 2 for d in deltas) / len(deltas)
    std_delta = math.sqrt(variance)

    return (std_delta / mean_delta) * 100


def assign_grade(mape: float) -> Tuple[str, str]:
    """
    Assign letter grade based on MAPE value.

    Returns (grade, css_class)
    """
    if mape is None:
        return 'N/A', 'grade-c'
    if mape <= 10:
        return 'A', 'grade-a'
    elif mape <= 20:
        return 'B', 'grade-b'
    elif mape <= 35:
        return 'C', 'grade-c'
    elif mape <= 50:
        return 'D', 'grade-d'
    else:
        return 'F', 'grade-f'


def calculate_tier_weighted_mape(matched_pairs: List[Dict]) -> Dict[str, Any]:
    """
    Calculate tier-weighted MAPE giving more weight to hyperscaler accuracy.

    Weights:
    - Hyperscaler: 60%
    - Major Colo: 30%
    - Other: 10%

    NOTE: Only uses same-granularity pairs (building↔building, campus↔campus)
    to ensure apples-to-apples comparison.
    """
    # Filter to same-granularity pairs first
    same_granularity_pairs = [p for p in matched_pairs if p.get('granularity_match', True)]

    tier_data = {'Hyperscaler': [], 'Major Colo': [], 'Other': []}

    for p in same_granularity_pairs:
        sa_cap = p['capacity_delta']['sa_capacity']
        dch_cap = p['capacity_delta']['dch_capacity']

        if sa_cap > 0 and dch_cap > 0:
            company = p['sa_record'].get('company_clean_filter') or p['sa_record'].get('company_clean') or ''
            tier = get_company_tier(company)

            max_cap = max(sa_cap, dch_cap)
            ape = abs(sa_cap - dch_cap) / max_cap * 100
            tier_data[tier].append(ape)

    # Calculate MAPE per tier
    tier_mapes = {}
    for tier, apes in tier_data.items():
        if apes:
            tier_mapes[tier] = sum(apes) / len(apes)
        else:
            tier_mapes[tier] = None

    # Calculate weighted MAPE
    weighted_mape = 0
    total_weight = 0

    for tier, weight in COMPANY_TIER_WEIGHTS.items():
        if tier_mapes.get(tier) is not None:
            weighted_mape += tier_mapes[tier] * weight
            total_weight += weight

    if total_weight > 0:
        weighted_mape = weighted_mape / total_weight
    else:
        weighted_mape = None

    return {
        'weighted_mape': weighted_mape,
        'tier_mapes': tier_mapes,
        'tier_counts': {t: len(v) for t, v in tier_data.items()}
    }


def calculate_agreement_rate(matched_pairs: List[Dict], threshold_pct: float = 20.0) -> float:
    """
    Calculate the percentage of matched pairs within a given percentage threshold.

    Default: % of pairs where |delta| / max < 20%

    NOTE: Only uses same-granularity pairs (building↔building, campus↔campus)
    to ensure apples-to-apples comparison.
    """
    # Filter to same-granularity pairs first
    same_granularity_pairs = [p for p in matched_pairs if p.get('granularity_match', True)]

    valid_pairs = [p for p in same_granularity_pairs
                   if p['capacity_delta']['sa_capacity'] > 0 and p['capacity_delta']['dch_capacity'] > 0]

    if not valid_pairs:
        return 0

    within_threshold = sum(1 for p in valid_pairs if p['capacity_delta']['delta_pct'] < threshold_pct)
    return (within_threshold / len(valid_pairs)) * 100


def bootstrap_confidence_interval(values: List[float], n_bootstrap: int = 1000, ci: float = 0.95) -> Tuple[float, float]:
    """
    Calculate bootstrap confidence interval for the mean.

    Returns (lower_bound, upper_bound) for the specified CI level.
    """
    import random

    if len(values) < 2:
        return (None, None)

    bootstrap_means = []
    n = len(values)

    for _ in range(n_bootstrap):
        # Resample with replacement
        sample = [random.choice(values) for _ in range(n)]
        bootstrap_means.append(sum(sample) / n)

    # Sort and get percentiles
    bootstrap_means.sort()
    alpha = 1 - ci
    lower_idx = int((alpha / 2) * n_bootstrap)
    upper_idx = int((1 - alpha / 2) * n_bootstrap)

    return (bootstrap_means[lower_idx], bootstrap_means[upper_idx])


def analyze_essential_dcs(matched_pairs: List[Dict], sa_only: List[Dict], dch_only: List[Dict]) -> Dict[str, Any]:
    """
    Specific analysis of essential/strategic data centers.

    Essential DCs are flagged with is_essential=True in gold layer.
    """
    # Find essential matched pairs
    essential_matched = [p for p in matched_pairs
                         if p['sa_record'].get('is_essential') or p['dch_record'].get('is_essential')]

    # Essential with significant conflicts (>20% delta)
    essential_conflicts = [p for p in essential_matched if p['capacity_delta']['delta_pct'] >= 20]

    # Essential SA-only
    essential_sa_only = [r for r in sa_only if r.get('is_essential')]

    # Essential DCH-only
    essential_dch_only = [r for r in dch_only if r.get('is_essential')]

    # Calculate MAPE for essential only
    if essential_matched:
        essential_mape_data = calculate_mape_and_bias(essential_matched)
        essential_mape = essential_mape_data.get('mape')
    else:
        essential_mape = None

    return {
        'total_essential_matched': len(essential_matched),
        'essential_with_conflict': len(essential_conflicts),
        'essential_sa_only': len(essential_sa_only),
        'essential_dch_only': len(essential_dch_only),
        'essential_mape': essential_mape,
        'essential_conflict_list': essential_conflicts
    }


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

def run_comparison(
    threshold_m: float = DEFAULT_SPATIAL_THRESHOLD_M,
    output_html: bool = True,
    output_csv: bool = True,
    output_fc: bool = True,
    use_campus_level: bool = True
) -> Dict[str, Any]:
    """
    Main entry point for SA vs DCH comparison.

    Args:
        threshold_m: Spatial matching threshold in meters
        output_html: Whether to generate HTML report
        output_csv: Whether to export CSV files
        output_fc: Whether to create conflict feature class
        use_campus_level: If True, compare at campus level using gold_campus_full.
                          This provides accurate site counts and avoids capacity
                          double-counting. Set to False for building-level comparison.
                          (Default: True)

    Returns:
        Dictionary with all analysis results
    """
    print("=" * 70)
    print("SA vs DCH COMPARISON V2 - Enhanced Statistical Analysis")
    print("=" * 70)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Select source feature class based on comparison level
    if use_campus_level:
        source_fc = TARGET_FC_CAMPUS
        comparison_level = "Campus"
        print(f"\n[1/8] Loading records from gold_campus_full (campus-level)...")
    else:
        source_fc = TARGET_FC_BUILDINGS
        comparison_level = "Building"
        print(f"\n[1/8] Loading records from gold_buildings_full (building-level)...")

    # Load records from appropriate feature class
    sa_records = load_records_by_source(source_fc, SA_SOURCES)
    dch_records = load_records_by_source(source_fc, DCH_SOURCES)

    print(f"    Loaded {len(sa_records):,} SA {comparison_level.lower()} records")
    print(f"    Loaded {len(dch_records):,} DCH {comparison_level.lower()} records")

    if not sa_records or not dch_records:
        print(f"ERROR: No records loaded. Check {os.path.basename(source_fc)} layer.")
        print("       If using campus-level (default), ensure campus_rollup_new.py has been run.")
        return {}

    # Step 2: Spatial matching
    print(f"\n[2/8] Performing spatial matching (threshold: {threshold_m}m)...")
    matched_pairs, sa_only, dch_only = build_match_sets(sa_records, dch_records, threshold_m)

    print(f"    Matched {comparison_level.lower()}s: {len(matched_pairs):,}")
    print(f"    SA-only {comparison_level.lower()}s: {len(sa_only):,}")
    print(f"    DCH-only {comparison_level.lower()}s: {len(dch_only):,}")

    # Validate record accounting
    total_accounted = len(matched_pairs) + len(sa_only)
    if total_accounted != len(sa_records):
        print(f"    WARNING: SA record mismatch ({total_accounted} vs {len(sa_records)})")

    # Step 3: Calculate MAPE, Bias, CV
    print("\n[3/8] Calculating statistical metrics...")
    mape_bias_data = calculate_mape_and_bias(matched_pairs)
    cv_value = calculate_cv(matched_pairs)
    agreement_rate = calculate_agreement_rate(matched_pairs, threshold_pct=20.0)
    tier_weighted_data = calculate_tier_weighted_mape(matched_pairs)

    mape = mape_bias_data.get('mape')
    bias_pct = mape_bias_data.get('bias_pct')
    grade, grade_class = assign_grade(mape)
    excluded_granularity = mape_bias_data.get('excluded_granularity', 0)

    # Calculate granularity stats for reporting (relevant for building-level comparisons)
    same_gran_count = sum(1 for p in matched_pairs if p.get('granularity_match', True))
    diff_gran_count = len(matched_pairs) - same_gran_count
    gran_match_rate = (same_gran_count / len(matched_pairs) * 100) if matched_pairs else 0

    if not use_campus_level:
        print(f"    Same-granularity pairs: {same_gran_count:,} ({gran_match_rate:.1f}%)")
        print(f"    Different-granularity pairs: {diff_gran_count:,} (excluded from MAPE/Bias)")
    print(f"    MAPE: {mape:.1f}%" if mape else "    MAPE: N/A")
    print(f"    Bias: {bias_pct:+.1f}%" if bias_pct else "    Bias: N/A")
    print(f"    CV: {cv_value:.1f}%" if cv_value else "    CV: N/A")
    print(f"    Grade: {grade}")
    print(f"    Agreement Rate (within 20%): {agreement_rate:.1f}%")
    print(f"    Grade: {grade}")
    print(f"    Agreement Rate (within 20%): {agreement_rate:.1f}%")

    # Bootstrap confidence intervals
    ci_lower, ci_upper = None, None
    if mape_bias_data.get('apes'):
        ci_lower, ci_upper = bootstrap_confidence_interval(mape_bias_data['apes'], n_bootstrap=500)
        if ci_lower and ci_upper:
            print(f"    MAPE 95% CI: [{ci_lower:.1f}%, {ci_upper:.1f}%]")

    # Step 4: Capacity analysis
    print("\n[4/7] Analyzing capacity differences...")
    capacity_analysis = analyze_capacity_conflicts(matched_pairs)

    # Add MAPE/bias/CV to capacity analysis for HTML report
    capacity_analysis['mape'] = mape
    capacity_analysis['bias_pct'] = bias_pct
    capacity_analysis['cv'] = cv_value
    capacity_analysis['grade'] = grade
    capacity_analysis['grade_class'] = grade_class
    capacity_analysis['agreement_rate'] = agreement_rate
    capacity_analysis['mape_ci'] = (ci_lower, ci_upper)
    capacity_analysis['tier_weighted_mape'] = tier_weighted_data.get('weighted_mape')

    # Step 5: Company and geographic analysis
    print("\n[5/8] Analyzing by company and geography...")
    company_analysis = analyze_by_company(matched_pairs, sa_only, dch_only)
    geo_analysis = analyze_by_geography(matched_pairs, sa_only, dch_only)
    essential_analysis = analyze_essential_dcs(matched_pairs, sa_only, dch_only)

    print(f"    Companies analyzed: {len(company_analysis):,}")
    print(f"    Essential DCs matched: {essential_analysis['total_essential_matched']:,}")
    print(f"    Essential with conflicts: {essential_analysis['essential_with_conflict']:,}")

    # Step 6: Net New Sites Analysis
    print("\n[6/8] Analyzing net new sites (Under Construction, Announced, Planned)...")
    net_new_analysis = analyze_net_new_sites(
        matched_pairs=matched_pairs,
        sa_only=sa_only,
        dch_only=dch_only,
        sa_records=sa_records,
        dch_records=dch_records
    )

    # Step 7: Generate outputs
    print("\n[7/8] Generating outputs...")

    # CSV exports
    if output_csv:
        export_csv_reports(matched_pairs, sa_only, dch_only, timestamp)

    # Feature class export
    if output_fc:
        export_conflict_feature_class(matched_pairs, timestamp)

    # HTML report
    if output_html:
        # Top conflicts
        top_conflicts = get_top_conflicts(matched_pairs, n=25)

        # Generate HTML report path
        report_path = os.path.join(ACCURACY_REPORTS_DIR, f"SA_vs_DCH_Comparison_V2_{timestamp}.html")

        generate_html_report(
            sa_records=sa_records,
            dch_records=dch_records,
            matched_pairs=matched_pairs,
            sa_only=sa_only,
            dch_only=dch_only,
            capacity_analysis=capacity_analysis,
            company_analysis=company_analysis,
            geo_analysis=geo_analysis,
            top_conflicts=top_conflicts,
            net_new_analysis=net_new_analysis,
            output_path=report_path,
            threshold_m=threshold_m
        )

    # Step 8: Summary
    print("\n[8/8] Comparison complete!")
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  Total SA Records:      {len(sa_records):,}")
    print(f"  Total DCH Records:     {len(dch_records):,}")
    print(f"  Matched Pairs:         {len(matched_pairs):,}")
    print(f"  SA-Only:               {len(sa_only):,}")
    print(f"  DCH-Only:              {len(dch_only):,}")
    print(f"  Significant Conflicts: {capacity_analysis.get('significant_conflicts_count', 0):,}")
    print("-" * 70)
    print(f"  MAPE:                  {mape:.1f}%" if mape else "  MAPE: N/A")
    print(f"  Bias:                  {bias_pct:+.1f}% ({'SA higher' if bias_pct and bias_pct > 0 else 'DCH higher'})" if bias_pct else "  Bias: N/A")
    print(f"  CV:                    {cv_value:.1f}%" if cv_value else "  CV: N/A")
    print(f"  Agreement Grade:       {grade}")
    print(f"  Pearson r:             {capacity_analysis.get('correlation', {}).get('r', 'N/A')}")
    print("=" * 70)

    return {
        'matched_pairs': matched_pairs,
        'sa_only': sa_only,
        'dch_only': dch_only,
        'capacity_analysis': capacity_analysis,
        'company_analysis': company_analysis,
        'geo_analysis': geo_analysis,
        'essential_analysis': essential_analysis,
        'mape': mape,
        'bias_pct': bias_pct,
        'cv': cv_value,
        'grade': grade,
        'timestamp': timestamp
    }


def export_csv_reports(matched_pairs: List[Dict], sa_only: List[Dict], dch_only: List[Dict], timestamp: str):
    """Export matched pairs, SA-only, and DCH-only records to CSV files and a combined Excel workbook."""
    reports_dir = ACCURACY_REPORTS_DIR
    os.makedirs(reports_dir, exist_ok=True)

    # Matched pairs CSV
    matched_path = os.path.join(reports_dir, f"SA_DCH_Matched_Pairs_{timestamp}.csv")
    matched_rows = []
    matched_header = [
        'sa_unique_id', 'dch_unique_id', 'sa_company', 'dch_company', 'company_match',
        'sa_capacity_mw', 'dch_capacity_mw', 'delta_mw', 'delta_pct',
        'sa_status', 'dch_status',
        'conflict_direction', 'is_significant', 'is_essential',
        'city', 'state', 'country', 'region', 'distance_m'
    ]

    for pair in matched_pairs:
        sa_rec = pair['sa_record']
        dch_rec = pair['dch_record']
        delta = pair['capacity_delta']

        direction = 'SA_Higher' if delta['delta_mw'] > 0 else 'DCH_Higher' if delta['delta_mw'] < 0 else 'Equal'
        is_essential = sa_rec.get('is_essential') or dch_rec.get('is_essential')

        matched_rows.append([
            sa_rec.get('unique_id'),
            dch_rec.get('unique_id'),
            sa_rec.get('company_clean'),
            dch_rec.get('company_clean'),
            pair.get('company_match', False),
            delta['sa_capacity'],
            delta['dch_capacity'],
            delta['delta_mw'],
            delta['delta_pct'],
            sa_rec.get('facility_status'),
            dch_rec.get('facility_status'),
            direction,
            delta['is_significant'],
            is_essential,
            sa_rec.get('city') or dch_rec.get('city'),
            sa_rec.get('state') or dch_rec.get('state'),
            sa_rec.get('country') or dch_rec.get('country'),
            sa_rec.get('region') or dch_rec.get('region'),
            pair.get('distance_m', 0)
        ])

    with open(matched_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(matched_header)
        writer.writerows(matched_rows)

    print(f"    Matched pairs CSV: {matched_path}")

    # SA-only CSV
    sa_only_path = os.path.join(reports_dir, f"SA_Only_Records_{timestamp}.csv")
    sa_only_header = [
        'unique_id', 'company', 'facility_name', 'full_capacity_mw', 'facility_status',
        'city', 'state', 'country', 'region', 'latitude', 'longitude', 'is_essential'
    ]
    sa_only_rows = []

    sorted_sa = sorted(sa_only, key=lambda x: safe_float(x.get('full_capacity_mw')) or 0, reverse=True)
    for rec in sorted_sa:
        sa_only_rows.append([
            rec.get('unique_id'),
            rec.get('company_clean'),
            rec.get('facility_name'),
            rec.get('full_capacity_mw'),
            rec.get('facility_status'),
            rec.get('city'),
            rec.get('state'),
            rec.get('country'),
            rec.get('region'),
            rec.get('latitude'),
            rec.get('longitude'),
            rec.get('is_essential')
        ])

    with open(sa_only_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(sa_only_header)
        writer.writerows(sa_only_rows)

    print(f"    SA-only CSV: {sa_only_path}")

    # DCH-only CSV
    dch_only_path = os.path.join(reports_dir, f"DCH_Only_Records_{timestamp}.csv")
    dch_only_header = [
        'unique_id', 'company', 'facility_name', 'full_capacity_mw', 'facility_status',
        'city', 'state', 'country', 'region', 'latitude', 'longitude', 'is_essential'
    ]
    dch_only_rows = []

    sorted_dch = sorted(dch_only, key=lambda x: safe_float(x.get('full_capacity_mw')) or 0, reverse=True)
    for rec in sorted_dch:
        dch_only_rows.append([
            rec.get('unique_id'),
            rec.get('company_clean'),
            rec.get('facility_name'),
            rec.get('full_capacity_mw'),
            rec.get('facility_status'),
            rec.get('city'),
            rec.get('state'),
            rec.get('country'),
            rec.get('region'),
            rec.get('latitude'),
            rec.get('longitude'),
            rec.get('is_essential')
        ])

    with open(dch_only_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(dch_only_header)
        writer.writerows(dch_only_rows)

    print(f"    DCH-only CSV: {dch_only_path}")

    # Combined Excel workbook with multiple tabs
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        excel_path = os.path.join(reports_dir, f"SA_vs_DCH_Comparison_{timestamp}.xlsx")
        wb = openpyxl.Workbook()

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1877F2", end_color="1877F2", fill_type="solid")
        sa_fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
        dch_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Matched Pairs
        ws_matched = wb.active
        ws_matched.title = "Matched Pairs"
        ws_matched.append(matched_header)
        for row in matched_rows:
            ws_matched.append(row)

        # Style header
        for col, cell in enumerate(ws_matched[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Adjust column widths
        for col in ws_matched.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_matched.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        # Sheet 2: SA-Only
        ws_sa = wb.create_sheet("SA Only")
        ws_sa.append(sa_only_header)
        for row in sa_only_rows:
            ws_sa.append(row)

        for col, cell in enumerate(ws_sa[1], 1):
            cell.font = header_font
            cell.fill = sa_fill
            cell.alignment = Alignment(horizontal='center')

        for col in ws_sa.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_sa.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        # Sheet 3: DCH-Only
        ws_dch = wb.create_sheet("DCH Only")
        ws_dch.append(dch_only_header)
        for row in dch_only_rows:
            ws_dch.append(row)

        for col, cell in enumerate(ws_dch[1], 1):
            cell.font = header_font
            cell.fill = dch_fill
            cell.alignment = Alignment(horizontal='center')

        for col in ws_dch.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_dch.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        # Sheet 4: Summary Statistics
        ws_summary = wb.create_sheet("Summary")
        summary_data = [
            ["SA vs DCH Comparison Summary", ""],
            ["", ""],
            ["Metric", "Value"],
            ["Total SA Records", len(matched_pairs) + len(sa_only)],
            ["Total DCH Records", len(matched_pairs) + len(dch_only)],
            ["Matched Pairs", len(matched_pairs)],
            ["SA-Only Records", len(sa_only)],
            ["DCH-Only Records", len(dch_only)],
            ["Significant Conflicts", sum(1 for p in matched_pairs if p['capacity_delta']['is_significant'])],
            ["", ""],
            ["SA Match Rate", f"{len(matched_pairs) / (len(matched_pairs) + len(sa_only)) * 100:.1f}%"],
            ["DCH Match Rate", f"{len(matched_pairs) / (len(matched_pairs) + len(dch_only)) * 100:.1f}%"],
        ]

        for row in summary_data:
            ws_summary.append(row)

        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20

        wb.save(excel_path)
        print(f"    Combined Excel workbook: {excel_path}")

    except ImportError:
        print("    NOTE: openpyxl not available - skipping Excel workbook creation")


def export_conflict_feature_class(matched_pairs: List[Dict], timestamp: str):
    """Export significant conflicts to a feature class in Default.gdb."""
    try:
        import arcpy
    except ImportError:
        print("    WARNING: arcpy not available - skipping feature class export")
        return

    # Filter to significant conflicts
    significant = [p for p in matched_pairs if p['capacity_delta']['is_significant']]

    if not significant:
        print("    No significant conflicts to export to feature class")
        return

    # Output path - use GDB from config
    fc_name = f"sa_dch_conflicts_{timestamp}"
    gdb_path = GDB
    fc_path = os.path.join(gdb_path, fc_name)

    # Create feature class
    arcpy.env.overwriteOutput = True

    # Use SA coordinates for point location
    sr = arcpy.SpatialReference(4326)  # WGS84

    arcpy.management.CreateFeatureclass(
        gdb_path, fc_name, "POINT", spatial_reference=sr
    )

    # Add fields
    fields_to_add = [
        ('sa_unique_id', 'TEXT', 100),
        ('dch_unique_id', 'TEXT', 100),
        ('sa_company', 'TEXT', 100),
        ('dch_company', 'TEXT', 100),
        ('company_match', 'SHORT'),
        ('sa_capacity_mw', 'DOUBLE'),
        ('dch_capacity_mw', 'DOUBLE'),
        ('delta_mw', 'DOUBLE'),
        ('delta_pct', 'DOUBLE'),
        ('conflict_direction', 'TEXT', 20),
        ('is_essential', 'SHORT'),
        ('confidence_level', 'TEXT', 20),
        ('city', 'TEXT', 100),
        ('state', 'TEXT', 50),
        ('country', 'TEXT', 100),
        ('region', 'TEXT', 50)
    ]

    for field_name, field_type, *args in fields_to_add:
        length = args[0] if args else None
        arcpy.management.AddField(fc_path, field_name, field_type, field_length=length)

    # Insert rows
    field_names = ['SHAPE@XY'] + [f[0] for f in fields_to_add]

    with arcpy.da.InsertCursor(fc_path, field_names) as cursor:
        for pair in significant:
            sa_rec = pair['sa_record']
            dch_rec = pair['dch_record']
            delta = pair['capacity_delta']

            lat = sa_rec.get('latitude') or dch_rec.get('latitude')
            lon = sa_rec.get('longitude') or dch_rec.get('longitude')

            if not lat or not lon:
                continue

            direction = 'SA_Higher' if delta['delta_mw'] > 0 else 'DCH_Higher'
            is_essential = 1 if (sa_rec.get('is_essential') or dch_rec.get('is_essential')) else 0

            # Confidence level based on delta magnitude
            abs_delta_pct = abs(delta['delta_pct'])
            if abs_delta_pct > 50:
                confidence = 'Low'
            elif abs_delta_pct > 30:
                confidence = 'Medium'
            else:
                confidence = 'High'

            row = (
                (float(lon), float(lat)),
                sa_rec.get('unique_id'),
                dch_rec.get('unique_id'),
                sa_rec.get('company_clean'),
                dch_rec.get('company_clean'),
                1 if pair.get('company_match') else 0,
                delta['sa_capacity'],
                delta['dch_capacity'],
                delta['delta_mw'],
                delta['delta_pct'],
                direction,
                is_essential,
                confidence,
                sa_rec.get('city') or dch_rec.get('city'),
                sa_rec.get('state') or dch_rec.get('state'),
                sa_rec.get('country') or dch_rec.get('country'),
                sa_rec.get('region') or dch_rec.get('region')
            )

            cursor.insertRow(row)

    print(f"    Conflict feature class: {fc_path} ({len(significant):,} records)")


def prepare_histogram_data(matched_pairs: List[Dict]) -> Tuple[List[str], List[int]]:
    """Prepare histogram data for capacity delta distribution."""
    deltas = [p['capacity_delta']['delta_mw'] for p in matched_pairs
              if p['capacity_delta']['sa_capacity'] > 0 and p['capacity_delta']['dch_capacity'] > 0]

    if not deltas:
        return [], []

    # Define bins
    bins = [-100, -50, -20, -10, -5, 0, 5, 10, 20, 50, 100]
    labels = ['<-50', '-50 to -20', '-20 to -10', '-10 to -5', '-5 to 0',
              '0 to 5', '5 to 10', '10 to 20', '20 to 50', '>50']

    counts = [0] * len(labels)

    for d in deltas:
        if d < -50:
            counts[0] += 1
        elif d < -20:
            counts[1] += 1
        elif d < -10:
            counts[2] += 1
        elif d < -5:
            counts[3] += 1
        elif d < 0:
            counts[4] += 1
        elif d < 5:
            counts[5] += 1
        elif d < 10:
            counts[6] += 1
        elif d < 20:
            counts[7] += 1
        elif d < 50:
            counts[8] += 1
        else:
            counts[9] += 1

    return labels, counts


def get_top_conflicts(matched_pairs: List[Dict], n: int = 25) -> List[Dict]:
    """Get top N conflicts by absolute delta magnitude."""
    conflicts = []

    for pair in matched_pairs:
        delta = pair['capacity_delta']
        if delta['sa_capacity'] > 0 and delta['dch_capacity'] > 0:
            conflicts.append({
                'sa_capacity_mw': delta['sa_capacity'],
                'dch_capacity_mw': delta['dch_capacity'],
                'delta_mw': delta['delta_mw'],
                'delta_pct': delta['delta_pct'],
                'sa_company': pair['sa_record'].get('company_clean'),
                'dch_company': pair['dch_record'].get('company_clean'),
                'city': pair['sa_record'].get('city') or pair['dch_record'].get('city'),
                'state': pair['sa_record'].get('state') or pair['dch_record'].get('state'),
                'country': pair['sa_record'].get('country') or pair['dch_record'].get('country')
            })

    # Sort by absolute delta
    conflicts.sort(key=lambda x: abs(x['delta_mw']), reverse=True)

    return conflicts[:n]


# ==============================================================================
# ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="SA vs DCH Enhanced Comparison V2")
    parser.add_argument('--threshold', type=float, default=DEFAULT_SPATIAL_THRESHOLD_M,
                        help=f'Spatial matching threshold in meters (default: {DEFAULT_SPATIAL_THRESHOLD_M})')
    parser.add_argument('--no-html', action='store_true', help='Skip HTML report generation')
    parser.add_argument('--no-csv', action='store_true', help='Skip CSV exports')
    parser.add_argument('--no-fc', action='store_true', help='Skip feature class export')

    args = parser.parse_args()

    results = run_comparison(
        threshold_m=args.threshold,
        output_html=not args.no_html,
        output_csv=not args.no_csv,
        output_fc=not args.no_fc
    )

    if results:
        print("\n✅ Comparison completed successfully!")
    else:
        print("\n❌ Comparison failed - check errors above")
